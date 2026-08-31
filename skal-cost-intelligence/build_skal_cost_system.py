# -*- coding: utf-8 -*-
"""
SKAL INDUSTRIAL COST INTELLIGENCE SYSTEM  —  gerador do workbook
================================================================
Gera um arquivo Excel profissional, modular, auditavel e parametrizado
para inteligencia de custos industriais da SKAL / Grupo Kalfix.

Filosofia:
  - Este script E a fonte da verdade da estrutura. Rode-o para reconstruir
    o workbook do zero. Nenhum numero critico e "magico": premissas ficam
    em 01_CONTROL_PANEL e sao referenciadas por formula.
  - Onde nao ha dado real da SKAL, criamos o CAMPO DE ENTRADA (input) e
    registramos a lacuna em DATA_GAPS. Linhas de exemplo sao marcadas
    explicitamente como "DEMO" para permitir os testes matematicos de
    reconciliacao (Standard + Variances = Actual, BOM fecha, etc.) sem
    apresentar numero inventado como se fosse dado oficial da empresa.
  - Ligar/desligar o dataset de demonstracao: CONTROL_PANEL!MODO_DEMO.

Uso:  python3 build_skal_cost_system.py
Saida: SKAL_Industrial_Cost_Intelligence_System.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT = "SKAL_Industrial_Cost_Intelligence_System.xlsx"

# ---------------------------------------------------------------------------
# PALETA / ESTILOS  (convencao: INPUT amarelo, FORMULA branco, LINKED azul,
# CONTROL verde, ALERT vermelho, KEY cinza, HEADER navy)
# ---------------------------------------------------------------------------
NAVY   = "1F3864"
BLUE   = "2E5496"
INK    = "1F4E78"
GREY_D = "808080"

FILL_HEADER  = PatternFill("solid", fgColor=NAVY)
FILL_SUB     = PatternFill("solid", fgColor=BLUE)
FILL_INPUT   = PatternFill("solid", fgColor="FFF2CC")   # amarelo palido
FILL_LINKED  = PatternFill("solid", fgColor="DDEBF7")   # azul palido
FILL_CONTROL = PatternFill("solid", fgColor="E2EFDA")   # verde palido
FILL_ALERT   = PatternFill("solid", fgColor="F8CBAD")   # vermelho palido
FILL_KEY     = PatternFill("solid", fgColor="D9D9D9")   # cinza
FILL_BANNER  = PatternFill("solid", fgColor="FBE5D6")   # laranja claro (demo)
FILL_ZEBRA   = PatternFill("solid", fgColor="F2F6FB")

F_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_TITLE  = Font(name="Calibri", size=16, bold=True, color=NAVY)
F_SUB    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_SECT   = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_INPUT  = Font(name="Calibri", size=10, color="0000CC")
F_FORM   = Font(name="Calibri", size=10, color="000000")
F_LINK   = Font(name="Calibri", size=10, color=INK)
F_KEY    = Font(name="Calibri", size=10, bold=True, color="000000")
F_NOTE   = Font(name="Calibri", size=9, italic=True, color=GREY_D)
F_BOLD   = Font(name="Calibri", size=10, bold=True, color="000000")

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM = Border(bottom=Side(style="medium", color=NAVY))

AL_L = Alignment(horizontal="left",  vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_R = Alignment(horizontal="right",  vertical="center")

MONEY = '#,##0.00'
MONEY4 = '#,##0.0000'
NUM = '#,##0.00'
INT = '#,##0'
PCT = '0.0%'
PCT2 = '0.00%'

wb = openpyxl.Workbook()
wb.remove(wb.active)

# registry to build a table of contents
TOC = []   # (sheet, group, purpose)


def sheet(name, group, purpose, tab=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab
    TOC.append((name, group, purpose))
    return ws


def title(ws, text, sub=None, span=12):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text); c.font = F_TITLE; c.alignment = AL_L
    ws.row_dimensions[1].height = 26
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(2, 1, sub); s.font = F_NOTE; s.alignment = AL_L


def header_row(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row, start + i, h)
        c.font = F_HEADER; c.fill = FILL_HEADER; c.alignment = AL_C
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 30


def put(ws, r, c, val, *, style="form", fmt=None, bold=False, align=None,
        border=True, note=None):
    cell = ws.cell(r, c, val)
    fills = {"input": FILL_INPUT, "link": FILL_LINKED, "control": FILL_CONTROL,
             "alert": FILL_ALERT, "key": FILL_KEY, "form": None, "banner": FILL_BANNER}
    fonts = {"input": F_INPUT, "link": F_LINK, "control": F_BOLD, "alert": F_BOLD,
             "key": F_KEY, "form": F_FORM, "note": F_NOTE, "banner": F_BOLD}
    if style in fills and fills[style]:
        cell.fill = fills[style]
    cell.font = fonts.get(style, F_FORM)
    if bold:
        cell.font = Font(name="Calibri", size=10, bold=True,
                         color=cell.font.color)
    if fmt:
        cell.number_format = fmt
    if align == "l":
        cell.alignment = AL_L
    elif align == "r":
        cell.alignment = AL_R
    elif align == "c":
        cell.alignment = AL_C
    else:
        cell.alignment = AL_L
    if border:
        cell.border = BORDER_ALL
    if note:
        cell.comment = Comment(note, "SKAL Cost System")
    return cell


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def legend(ws, row, span=12):
    items = [("INPUT (editavel)", FILL_INPUT), ("FORMULA", None),
             ("LINKED (de outra aba)", FILL_LINKED),
             ("CONTROL", FILL_CONTROL), ("ALERT", FILL_ALERT),
             ("CHAVE/ID", FILL_KEY)]
    c = 1
    put(ws, row, c, "Legenda:", style="note"); c += 1
    for txt, fill in items:
        cell = ws.cell(row, c, txt)
        cell.font = F_NOTE
        if fill:
            cell.fill = fill
        cell.border = BORDER_ALL
        cell.alignment = AL_C
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c+1)
        c += 2


# ===========================================================================
# GENERIC TABLE SHEET  (master data / transactions with governance columns)
# ===========================================================================
# A "col" is a dict: name, w(idth), style(input/form/link/key), fmt, note
GOV = [  # governance columns appended to every critical master table
    {"name": "Fonte", "w": 16, "style": "input"},
    {"name": "Responsavel", "w": 16, "style": "input"},
    {"name": "Vigencia ini", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
    {"name": "Vigencia fim", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
    {"name": "Atualizado em", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
    {"name": "Comentario", "w": 26, "style": "input"},
]


def table_sheet(name, group, purpose, tsub, cols, demo=None, gov=False,
                tab=None, nrows=60, span=None):
    ws = sheet(name, group, purpose, tab=tab)
    allcols = list(cols) + (GOV if gov else [])
    span = span or max(12, len(allcols))
    title(ws, name.split("_", 1)[1].replace("_", " ") if "_" in name else name,
          tsub, span=span)
    legend(ws, 4, span=span)
    hrow = 6
    header_row(ws, hrow, [c["name"] for c in allcols])
    for i, c in enumerate(allcols):
        widths(ws, {get_column_letter(1 + i): c.get("w", 14)})
    # demo + blank input rows
    demo = demo or []
    for r_off in range(nrows):
        r = hrow + 1 + r_off
        drow = demo[r_off] if r_off < len(demo) else None
        for i, c in enumerate(allcols):
            col = 1 + i
            val = drow[i] if (drow is not None and i < len(drow)) else None
            st = c.get("style", "input")
            # formula cells (string starting with =) rendered as form
            if isinstance(val, str) and val.startswith("="):
                cell = put(ws, r, col, val, style="form", fmt=c.get("fmt"))
            else:
                cell = put(ws, r, col, val, style=st, fmt=c.get("fmt"))
            if r_off % 2 == 1 and st != "key" and not (isinstance(val, str) and val.startswith("=")):
                pass
        ws.row_dimensions[r].height = 15
    ws.freeze_panes = ws.cell(hrow + 1, 1)
    if demo:
        put(ws, hrow + 1 + len(demo) + 1, 1,
            "^ Linhas acima = dados DEMO (marcados). Abaixo: campos de entrada em branco para dados reais.",
            style="note", border=False)
    return ws


# ===========================================================================
# 00_README
# ===========================================================================
def build_readme():
    ws = sheet("00_README", "Doc", "Documentacao, metodologia e indice", tab="1F3864")
    widths(ws, {"A": 3, "B": 34, "C": 30, "D": 60, "E": 14})
    title(ws, "SKAL INDUSTRIAL COST INTELLIGENCE SYSTEM",
          "Digital Cost Twin da operacao industrial — Grupo Kalfix / SKAL. "
          "Gerado por build_skal_cost_system.py (fonte da verdade da estrutura).", span=5)
    r = 4
    blocks = [
        ("O QUE E", "Sistema modular e auditavel que responde, para qualquer SKU: quanto custou, "
         "quanto deveria ter custado, quanto custaria hoje, qual o desvio, onde/porque ocorreu, "
         "quanto e controlavel, e quanto vale elimina-lo."),
        ("CINCO FUNCOES", "1) Cost Accounting (quanto custou) 2) Cost Engineering (quanto deveria) "
         "3) Cost Control (onde desviou) 4) Cost Reduction (quanto economizar) "
         "5) Profitability Intelligence (onde ganha/perde dinheiro)."),
        ("COMO USAR", "Preencha as celulas AMARELAS (INPUT). Celulas brancas sao FORMULA; azuis sao "
         "dados de outra aba (LINKED); verdes sao CONTROL; vermelhas sao ALERT. Nao altere formulas "
         "estruturais. Premissas globais ficam em 01_CONTROL_PANEL."),
        ("MODO DEMO", "01_CONTROL_PANEL!MODO_DEMO liga um pequeno dataset de DEMONSTRACAO (marcado em "
         "cada linha) que permite os testes matematicos. Numeros DEMO NAO sao dados reais da SKAL — "
         "veja DATA_GAPS para o que precisa ser preenchido."),
        ("RECONCILIACAO", "Regra central: Standard Cost + Variances (PPV+Uso+Mix+Yield) = Actual Cost, "
         "residual -> 0. BOM fecha fisicamente. Nenhum custo e rateado sem driver causal "
         "(vai para UNALLOCATED)."),
    ]
    for h, txt in blocks:
        c = ws.cell(r, 2, h); c.font = F_SECT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        t = ws.cell(r, 3, txt); t.font = F_FORM; t.alignment = AL_L
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1
    # dictionary of indicators
    c = ws.cell(r, 2, "DICIONARIO DE INDICADORES-CHAVE"); c.font = F_SECT; r += 1
    header_row(ws, r, ["Indicador", "Definicao / Formula", "Unidade", "Resp."], start=2)
    r += 1
    inds = [
        ("Landed Net Cost", "Bruto - descontos - creditos recup. + frete/seguro/descarga + trib. nao recup.", "R$/t", "Compras/Fiscal"),
        ("Material Cost / Good Ton", "Custo total materiais consumidos / toneladas boas liberadas", "R$/t", "Producao/Contr."),
        ("PPV", "(Preco real - preco padrao) x quantidade real", "R$", "Compras"),
        ("Usage Variance", "(Qtd real - qtd padrao permitida) x preco padrao", "R$", "Producao"),
        ("Mix Variance", "Efeito da alteracao da composicao relativa da formula (sub-split do uso)", "R$", "Tecnico"),
        ("Yield Variance", "(Output bom esperado - obtido) valorizado ao custo padrao", "R$", "Producao"),
        ("Unused Capacity Cost", "Horas ociosas x custo/hora de capacidade pratica (NAO rateado ao produto)", "R$", "Controladoria"),
        ("COPQ", "Custo da ma qualidade: scrap+reprocesso+retrabalho+devolucao+frete reverso...", "R$", "Qualidade"),
        ("Cost Gap", "Actual -> Current Std -> Best Demonstrated -> Entitlement -> Teorico", "R$/t", "Controladoria"),
        ("Pocket Price", "Preco tabela - descontos - bonif - rebate - comissao - frete - custo prazo - inadimpl.", "R$", "Comercial"),
        ("Pocket Margin", "Pocket Price - custo reposicao - cost to serve - tributacao economica", "R$", "Comercial/Contr."),
        ("MC/hora do gargalo", "Margem de contribuicao por hora do recurso restritivo (TOC)", "R$/h", "Producao/Contr."),
        ("Saving Realizado", "Saving com evidencia operacional + confirmacao financeira + reconciliacao", "R$", "Controladoria"),
    ]
    for row in inds:
        for i, v in enumerate(row):
            put(ws, r, 2 + i, v, style="form", align="l")
        r += 1
    r += 1
    c = ws.cell(r, 2, "INDICE DE ABAS"); c.font = F_SECT; r += 1
    header_row(ws, r, ["Aba", "Grupo", "Proposito"], start=2); r += 1
    ws._toc_start = r  # filled after all sheets built
    ws.freeze_panes = "A4"
    return ws


# ===========================================================================
# 01_CONTROL_PANEL  (premissas globais + defined names)
# ===========================================================================
def build_control_panel():
    ws = sheet("01_CONTROL_PANEL", "Core", "Premissas globais parametrizadas", tab="C00000")
    widths(ws, {"A": 3, "B": 34, "C": 16, "D": 12, "E": 40, "F": 18, "G": 16})
    title(ws, "PAINEL DE CONTROLE — PREMISSAS GLOBAIS",
          "Toda premissa critica vive aqui e e referenciada por NOME. Nenhuma aliquota "
          "digitada dentro de formula. Edite apenas a coluna VALOR (amarela).", span=7)
    legend(ws, 4, span=7)
    r = 6
    header_row(ws, r, ["#", "Parametro", "Valor", "Un.", "Descricao / regra", "Fonte", "Responsavel"], start=1)
    r += 1
    # (defined_name, label, value, unit, fmt, desc, fonte, resp)  value None => DADO NECESSARIO
    P = [
        ("MODO_DEMO", "Modo demonstracao (TRUE/FALSE)", True, "-", "GENERAL",
         "TRUE liga dataset DEMO p/ testes; FALSE zera exemplos", "Controladoria", "Controller"),
        (None, "— CONVERSOES DE EMBALAGEM —", None, "", "", "", "", ""),
        ("KG_SACO_STD", "kg por saco (padrao)", 15, "kg", INT, "Linhas gerais SKAL", "Producao", "Sup. Prod."),
        ("SACO_PALETE_STD", "sacos por palete (padrao)", 100, "un", INT, "Linhas gerais", "Producao", "Sup. Prod."),
        ("KG_SACO_EXT", "kg por saco (linha Extras)", 20, "kg", INT, "Multiuso/Graute/Reboco/Contrapiso", "Producao", "Sup. Prod."),
        ("SACO_PALETE_EXT", "sacos por palete (Extras)", 75, "un", INT, "Linha Extras", "Producao", "Sup. Prod."),
        ("KG_PALETE", "kg por palete (todas)", 1500, "kg", INT, "1.500 kg/palete em ambas as linhas", "Producao", "Sup. Prod."),
        ("CONCHA_M3", "m3 por concha (Komatsu WA200-5)", 2.1, "m3", NUM, "Elo com Painel do Forno (areia)", "Producao", "Sup. Prod."),
        (None, "— ICMS (beneficio 80%) —", None, "", "", "", "", ""),
        ("ICMS_NOMINAL", "ICMS aliquota nominal", 0.225, "%", PCT2, "Aliquota cheia", "Fiscal", "Cont. Fiscal"),
        ("ICMS_REDUCAO", "ICMS reducao do beneficio", 0.80, "%", PCT2, "Reducao de 80% do ICMS devido", "Fiscal", "Cont. Fiscal"),
        ("ICMS_PCT_DEVIDO", "ICMS % efetivamente devido", "=1-ICMS_REDUCAO", "%", PCT2, "1 menos reducao (da 20%)", "Fiscal", "Cont. Fiscal"),
        ("ICMS_EFETIVO", "ICMS efetivo (calc.)", "=ICMS_NOMINAL*ICMS_PCT_DEVIDO", "%", PCT2, "nominal x %devido (=4,5%) NUNCA hardcoded", "Fiscal", "Cont. Fiscal"),
        (None, "— PIS/COFINS (Lucro Real, nao cumulativo) —", None, "", "", "", "", ""),
        ("PIS_DEBITO", "PIS debito", 0.0165, "%", PCT2, "Nao cumulativo", "Fiscal", "Cont. Fiscal"),
        ("COFINS_DEBITO", "COFINS debito", 0.076, "%", PCT2, "Nao cumulativo (soma 9,25%)", "Fiscal", "Cont. Fiscal"),
        (None, "— IRPJ / CSLL / SUDENE —", None, "", "", "", "", ""),
        ("IRPJ_ALIQ", "IRPJ aliquota", 0.15, "%", PCT2, "15% + adicional", "Fiscal", "Cont. Fiscal"),
        ("IRPJ_ADIC", "IRPJ adicional (>20k/mes)", 0.10, "%", PCT2, "10% sobre excedente", "Fiscal", "Cont. Fiscal"),
        ("CSLL_ALIQ", "CSLL aliquota", 0.09, "%", PCT2, "NUNCA aplicar reducao SUDENE a CSLL", "Fiscal", "Cont. Fiscal"),
        ("SUDENE_REDUCAO", "SUDENE reducao IRPJ elegivel", 0.75, "%", PCT2, "75% do IRPJ elegivel (parametrizavel)", "Fiscal", "Diretoria"),
        (None, "— FINANCEIRO —", None, "", "", "", "", ""),
        ("CUSTO_CAPITAL_ANO", "Custo de capital anual", 0.18, "%", PCT2, "Para custo de prazo e estoque (DEMO)", "Financeiro", "CFO"),
        ("INADIMPL_PCT", "Inadimplencia esperada (%)", 0.015, "%", PCT2, "Taxa historica parametrizada (DEMO)", "Financeiro", "Credito"),
        ("DIAS_ANO", "Dias no ano (base financeira)", 365, "d", INT, "Base p/ custo de prazo", "Financeiro", "CFO"),
        (None, "— MATERIALIDADE / GOVERNANCA —", None, "", "", "", "", ""),
        ("MAT_CLASSE_A", "Materialidade Classe A (R$/ano)", 100000, "R$", INT, "Impacto alto", "Controladoria", "Controller"),
        ("MAT_CLASSE_B", "Materialidade Classe B (R$/ano)", 25000, "R$", INT, "Impacto medio", "Controladoria", "Controller"),
        ("VOLUME_ANO_T", "Volume anual referencia (t)", 70000, "t", INT, "Para anualizar gaps (DEMO)", "Comercial", "Diretoria"),
    ]
    for row in P:
        dn, label, val, unit, fmt, desc, fonte, resp = row
        if val is None and dn is None:
            # section subheader
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c = ws.cell(r, 2, label); c.font = F_SUB; c.fill = FILL_SUB
            c.alignment = AL_L
            put(ws, r, 1, "", style="form")
            r += 1
            continue
        put(ws, r, 1, dn or "", style="key", align="c")
        put(ws, r, 2, label, style="form", align="l")
        # value cell
        if isinstance(val, str) and val.startswith("="):
            vc = put(ws, r, 3, val, style="form", fmt=(fmt if fmt in (PCT2, PCT, NUM, INT, MONEY) else None), align="r")
        elif val is None:
            vc = put(ws, r, 3, "DADO NECESSARIO", style="alert", align="c")
        else:
            vc = put(ws, r, 3, val, style="input", fmt=(fmt if fmt in (PCT2, PCT, NUM, INT, MONEY) else None), align="r")
        put(ws, r, 4, unit, style="form", align="c")
        put(ws, r, 5, desc, style="form", align="l")
        put(ws, r, 6, fonte, style="input", align="l")
        put(ws, r, 7, resp, style="input", align="l")
        if dn:
            try:
                wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
                    dn, attr_text="'01_CONTROL_PANEL'!$C$%d" % r))
            except Exception as e:
                print("defname", dn, e)
        r += 1
    ws.freeze_panes = "A7"
    # boolean validation on MODO_DEMO handled implicitly
    return ws


# ===========================================================================
# 02_CALENDAR
# ===========================================================================
def build_calendar():
    cols = [
        {"name": "PERIOD_ID", "w": 12, "style": "key"},
        {"name": "Ano", "w": 8, "style": "input", "fmt": INT},
        {"name": "Mes", "w": 8, "style": "input", "fmt": INT},
        {"name": "Data ini", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
        {"name": "Data fim", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
        {"name": "Dias uteis", "w": 10, "style": "input", "fmt": INT},
        {"name": "Regime tributario vigente", "w": 26, "style": "input"},
        {"name": "Status", "w": 12, "style": "input"},
    ]
    demo = [
        ["2026-01", 2026, 1, "2026-01-01", "2026-01-31", 22, "ICMS benef. 80% + PIS/COFINS n-cum", "Fechado"],
        ["2026-02", 2026, 2, "2026-02-01", "2026-02-28", 20, "ICMS benef. 80% + PIS/COFINS n-cum", "Fechado"],
        ["2026-03", 2026, 3, "2026-03-01", "2026-03-31", 21, "ICMS benef. 80% + PIS/COFINS n-cum", "Aberto"],
    ]
    table_sheet("02_CALENDAR", "Core", "Periodos e vigencias",
                "Controle temporal. A reforma tributaria e tratada por DATA DE VIGENCIA "
                "(ver 10_TAX_RULES), nunca sobrescrevendo regra antiga.", cols, demo=demo)


# ---- shared header for custom computational sheets --------------------------
def custom_head(name, group, purpose, tsub, headers, wspec, tab=None, span=None):
    ws = sheet(name, group, purpose, tab=tab)
    span = span or max(12, len(headers))
    title(ws, name.split("_", 1)[1].replace("_", " "), tsub, span=span)
    legend(ws, 4, span=span)
    header_row(ws, 6, headers)
    widths(ws, wspec)
    ws.freeze_panes = "A7"
    return ws


# demo dimensions (row 7 = first data row on every table)
R0 = 7  # first data row

# ===========================================================================
# 03_PRODUCTS
# ===========================================================================
def build_products():
    H = ["PRODUCT_ID", "Descricao", "Familia", "Linha", "Unid.\ncomercial",
         "Extras?", "kg/saco", "sacos/\npalete", "kg/palete", "FORMULA_ID",
         "Versao", "ROUTING_ID", "RESOURCE\nprincipal", "Familia\nsetup",
         "NCM", "Unid.\nindustrial", "Status", "Vigencia\nini", "Vigencia\nfim"]
    W = {"A": 12, "B": 26, "C": 14, "D": 12, "E": 10, "F": 9, "G": 8, "H": 9,
         "I": 10, "J": 12, "K": 8, "L": 12, "M": 12, "N": 12, "O": 12, "P": 10,
         "Q": 10, "R": 12, "S": 12}
    ws = custom_head("03_PRODUCTS", "Master Data", "Cadastro de produtos (SKU)",
                     "Conversoes de embalagem automaticas. Excecao linha Extras "
                     "(Multiuso/Graute/Reboco/Contrapiso): 20 kg/saco, 75 sacos/palete.",
                     H, W, tab="1F3864", span=19)
    demo = [
        ["P001", "Multiuso", "Extras", "Extras", "saco", True, None, None, None,
         "F-P001", "v1", "RT-EXT", "RES-MIST1", "Extras", "3824.50.00", "t", "Ativo", "2026-01-01", ""],
        ["P002", "Argamassa Assentamento AC-III", "Assentamento", "Colante", "saco", False, None, None, None,
         "F-P002", "v1", "RT-ARG", "RES-MIST1", "Colante", "3824.50.00", "t", "Ativo", "2026-01-01", ""],
        ["P003", "Rejunte", "Rejunte", "Rejunte", "saco", False, None, None, None,
         "F-P003", "v1", "RT-REJ", "RES-MIST2", "Rejunte", "3824.50.00", "t", "Ativo", "2026-01-01", ""],
    ]
    r = R0
    for d in demo:
        for i, v in enumerate(d):
            col = i + 1
            if col in (7, 8, 9):
                continue
            st = "key" if col == 1 else "input"
            put(ws, r, col, v, style=st, align="c" if col in (1, 6, 11) else "l")
        # formulas: kg/saco, sacos/palete, kg/palete
        put(ws, r, 7, "=IF(F%d,KG_SACO_EXT,KG_SACO_STD)" % r, style="form", fmt=INT, align="c")
        put(ws, r, 8, "=IF(F%d,SACO_PALETE_EXT,SACO_PALETE_STD)" % r, style="form", fmt=INT, align="c")
        put(ws, r, 9, "=G%d*H%d" % (r, r), style="form", fmt=INT, align="c")
        r += 1
    # blank input rows
    for r in range(R0 + len(demo), R0 + 40):
        put(ws, r, 1, None, style="key")
        for col in range(2, 20):
            if col in (7, 8, 9):
                put(ws, r, col, "=IF($F%d=TRUE,%s,%s)" % (
                    r, "KG_SACO_EXT" if col == 7 else ("SACO_PALETE_EXT" if col == 8 else "G%d*H%d" % (r, r)),
                    "KG_SACO_STD" if col == 7 else "SACO_PALETE_STD"), style="form", fmt=INT, align="c") if col != 9 else \
                    put(ws, r, 9, "=G%d*H%d" % (r, r), style="form", fmt=INT, align="c")
            else:
                put(ws, r, col, None, style="input")
    return ws


# ===========================================================================
# 04_MATERIALS
# ===========================================================================
def build_materials():
    cols = [
        {"name": "MATERIAL_ID", "w": 12, "style": "key"},
        {"name": "Descricao", "w": 26, "style": "input"},
        {"name": "Grupo", "w": 14, "style": "input"},
        {"name": "Unid.", "w": 8, "style": "input"},
        {"name": "NCM", "w": 12, "style": "input"},
        {"name": "Funcao tecnica", "w": 20, "style": "input"},
        {"name": "Substituto (MATERIAL_ID)", "w": 16, "style": "input"},
        {"name": "Recuperavel ICMS?", "w": 12, "style": "input"},
        {"name": "Recuperavel PIS/COFINS?", "w": 14, "style": "input"},
        {"name": "Classe (A/B/C)", "w": 10, "style": "input"},
    ]
    demo = [
        ["M001", "Cimento CP-II-32", "Aglomerante", "t", "2523.29.10", "Aglomerante estrutural", "", "Sim", "Sim", "A"],
        ["M002", "Areia industrial seca", "Agregado", "t", "2505.10.00", "Agregado/carga", "", "Sim", "Sim", "A"],
        ["M003", "Cal hidratada CH-I", "Aglomerante", "t", "2522.20.00", "Plastificante/retencao", "", "Sim", "Sim", "B"],
        ["M004", "Aditivo retentor de agua (HEC)", "Aditivo", "kg", "3912.39.00", "Retencao de agua/trabalhab.", "", "Sim", "Sim", "B"],
        ["M005", "Pigmento oxido de ferro", "Aditivo", "kg", "2821.10.00", "Coloracao", "", "Sim", "Sim", "C"],
    ]
    table_sheet("04_MATERIALS", "Master Data", "Cadastro de materias-primas",
                "Cada MP tem chave unica, funcao tecnica e substituto. Recuperabilidade "
                "tributaria alimenta o Landed Net Cost.", cols, demo=demo, gov=True, tab="1F3864")


# ===========================================================================
# 05_SUPPLIERS
# ===========================================================================
def build_suppliers():
    cols = [
        {"name": "SUPPLIER_ID", "w": 12, "style": "key"},
        {"name": "Razao social", "w": 26, "style": "input"},
        {"name": "UF", "w": 6, "style": "input"},
        {"name": "Fornece (grupo)", "w": 16, "style": "input"},
        {"name": "Lead time (d)", "w": 10, "style": "input", "fmt": INT},
        {"name": "Lote min (t)", "w": 10, "style": "input", "fmt": NUM},
        {"name": "Condicao pgto (d)", "w": 12, "style": "input", "fmt": INT},
        {"name": "Rating qualidade", "w": 12, "style": "input"},
        {"name": "Estrategico?", "w": 10, "style": "input"},
    ]
    demo = [
        ["S001", "Cimento Fornecedor A", "CE", "Cimento", 3, 27.0, 28, "A", "Sim"],
        ["S002", "Mineracao Areia (ASSIS/SKAL)", "CE", "Areia", 2, 30.0, 21, "B", "Sim"],
        ["S003", "Quimica Aditivos B", "SP", "Aditivo/Pigmento", 10, 1.0, 30, "A", "Nao"],
    ]
    table_sheet("05_SUPPLIERS", "Master Data", "Cadastro de fornecedores",
                "Base para Landed Cost, Supplier TCO e Should Cost.", cols, demo=demo,
                gov=True, tab="1F3864")


# ===========================================================================
# 09_LANDED_COST  (custo liquido posto fabrica) — nucleo do custo de MP
# ===========================================================================
def build_landed_cost():
    H = ["MATERIAL_ID", "SUPPLIER_ID", "PERIOD", "Unid",
         "Preco bruto\n(R$/un)", "(-) Desconto\nincond.", "(-) Rebate\natribuivel",
         "(+) Frete\ninbound", "(+) Seguro", "(+) Descarga", "(+) Armazenag.",
         "(+) Trib. nao\nrecuperavel", "(-) Creditos\nrecuperaveis",
         "LANDED NET\n(R$/un)", "Fator\nun->t", "LANDED NET\n(R$/t)",
         "Desembolso\nfinanceiro", "Credito\nfiscal", "Custo\ncontabil"]
    W = {"A": 12, "B": 12, "C": 10, "D": 6}
    for col in "EFGHIJKLMNOPQRS":
        W[col] = 11
    ws = custom_head("09_LANDED_COST", "Procurement", "Custo liquido posto fabrica",
                     "LANDED NET = bruto - desconto - rebate + frete + seguro + descarga + "
                     "armazenagem + trib.nao-recup - creditos recuperaveis. Nunca usar preco de NF puro.",
                     H, W, tab="C55A11", span=19)
    # demo: R$/un values already at per-t basis for t materials; per-kg for kg materials
    demo = [
        # id, sup, period, unid, bruto, desc, rebate, frete, seguro, descarga, armaz, trib_nr, cred_rec
        ["M001", "S001", "2026-03", "t", 700, 20, 10, 45, 3, 8, 6, 0, 92],
        ["M002", "S002", "2026-03", "t", 55, 0, 0, 22, 1, 5, 3, 0, 20],
        ["M003", "S001", "2026-03", "t", 520, 15, 0, 40, 2, 6, 4, 0, 72],
        ["M004", "S003", "2026-03", "kg", 19.0, 0.4, 0, 0.6, 0.05, 0.1, 0.05, 0, 2.4],
        ["M005", "S003", "2026-03", "kg", 9.5, 0.2, 0, 0.4, 0.03, 0.05, 0.03, 0, 1.2],
    ]
    r = R0
    for d in demo:
        put(ws, r, 1, d[0], style="key", align="c")
        put(ws, r, 2, d[1], style="input", align="c")
        put(ws, r, 3, d[2], style="input", align="c")
        put(ws, r, 4, d[3], style="input", align="c")
        for i, v in enumerate(d[4:]):  # cols 5..13 inputs
            put(ws, r, 5 + i, v, style="input", fmt=MONEY4, align="r")
        # LANDED NET/un
        put(ws, r, 14, "=E{r}-F{r}-G{r}+H{r}+I{r}+J{r}+K{r}+L{r}-M{r}".format(r=r),
            style="form", fmt=MONEY4, align="r")
        # fator un->t (t=1 ; kg=1000)
        put(ws, r, 15, '=IF(D{r}="kg",1000,1)'.format(r=r), style="form", fmt=INT, align="c")
        # LANDED NET R$/t
        put(ws, r, 16, "=N{r}*O{r}".format(r=r), style="form", fmt=MONEY, align="r")
        # desembolso financeiro = bruto-desc-rebate+frete+seguro+descarga+armaz+trib (o que sai do caixa, sem credito)
        put(ws, r, 17, "=E{r}-F{r}-G{r}+H{r}+I{r}+J{r}+K{r}+L{r}".format(r=r), style="form", fmt=MONEY4, align="r")
        put(ws, r, 18, "=M{r}".format(r=r), style="form", fmt=MONEY4, align="r")
        # custo contabil = landed net (economico p/ estoque)
        put(ws, r, 19, "=N{r}".format(r=r), style="form", fmt=MONEY4, align="r")
        r += 1
    for r in range(R0 + len(demo), R0 + 40):
        put(ws, r, 1, None, style="key")
        for c in range(2, 4):
            put(ws, r, c, None, style="input")
        put(ws, r, 4, None, style="input")
        for c in range(5, 14):
            put(ws, r, c, None, style="input", fmt=MONEY4)
        put(ws, r, 14, "=E{r}-F{r}-G{r}+H{r}+I{r}+J{r}+K{r}+L{r}-M{r}".format(r=r), style="form", fmt=MONEY4)
        put(ws, r, 15, '=IF(D{r}="kg",1000,1)'.format(r=r), style="form", fmt=INT)
        put(ws, r, 16, "=N{r}*O{r}".format(r=r), style="form", fmt=MONEY)
        put(ws, r, 17, "=E{r}-F{r}-G{r}+H{r}+I{r}+J{r}+K{r}+L{r}".format(r=r), style="form", fmt=MONEY4)
        put(ws, r, 18, "=M{r}".format(r=r), style="form", fmt=MONEY4)
        put(ws, r, 19, "=N{r}".format(r=r), style="form", fmt=MONEY4)
    return ws


LANDED_LAST = R0 + 40  # range end for lookups


# ===========================================================================
# 10_TAX_RULES  &  11_TAX_CREDITS
# ===========================================================================
def build_tax_rules():
    cols = [
        {"name": "RULE_ID", "w": 10, "style": "key"},
        {"name": "Tributo", "w": 10, "style": "input"},
        {"name": "UF", "w": 6, "style": "input"},
        {"name": "Operacao", "w": 14, "style": "input"},
        {"name": "NCM", "w": 12, "style": "input"},
        {"name": "CFOP", "w": 8, "style": "input"},
        {"name": "CST", "w": 8, "style": "input"},
        {"name": "Aliq. nominal", "w": 10, "style": "input", "fmt": PCT2},
        {"name": "Beneficio", "w": 16, "style": "input"},
        {"name": "% reducao", "w": 9, "style": "input", "fmt": PCT2},
        {"name": "Aliq. efetiva", "w": 10, "style": "form", "fmt": PCT2},
        {"name": "Gera credito?", "w": 10, "style": "input"},
        {"name": "Fundo adic.", "w": 10, "style": "input", "fmt": PCT2},
    ]
    demo = [
        ["TR001", "ICMS", "CE", "Saida interna", "3824.50.00", "5101", "51", 0.225, "Reducao 80%", 0.80, "=H7*(1-J7)", "Nao", 0],
        ["TR002", "ICMS", "CE", "Entrada MP", "2523.29.10", "1101", "00", 0.180, "Credito integral", 0.0, "=H8*(1-J8)", "Sim", 0],
        ["TR003", "PIS", "-", "Saida", "-", "-", "-", 0.0165, "Nao cumulativo", 0.0, "=H9*(1-J9)", "Sim", 0],
        ["TR004", "COFINS", "-", "Saida", "-", "-", "-", 0.076, "Nao cumulativo", 0.0, "=H10*(1-J10)", "Sim", 0],
        ["TR005", "CBS", "-", "Saida (2027+)", "-", "-", "-", 0.0, "Reforma - preencher", 0.0, "=H11*(1-J11)", "Sim", 0],
        ["TR006", "IBS", "-", "Saida (2029+)", "-", "-", "-", 0.0, "Reforma - preencher", 0.0, "=H12*(1-J12)", "Sim", 0],
    ]
    table_sheet("10_TAX_RULES", "Tax", "Motor tributario (regras por vigencia)",
                "Nenhuma aliquota digitada dentro de formula de custo. Cada regra tem vigencia; "
                "reforma tributaria (CBS/IBS 2026-2033) por DATA, nunca sobrescrevendo a antiga.",
                cols, demo=demo, gov=True, tab="C55A11", span=19)


def build_tax_credits():
    cols = [
        {"name": "MATERIAL_ID", "w": 12, "style": "key"},
        {"name": "PERIOD", "w": 10, "style": "input"},
        {"name": "Base (R$)", "w": 12, "style": "input", "fmt": MONEY},
        {"name": "Credito ICMS potencial", "w": 14, "style": "form", "fmt": MONEY},
        {"name": "Credito PIS/COFINS potencial", "w": 16, "style": "form", "fmt": MONEY},
        {"name": "Credito escriturado", "w": 14, "style": "input", "fmt": MONEY},
        {"name": "Credito utilizado", "w": 14, "style": "input", "fmt": MONEY},
        {"name": "Credito monetizado", "w": 14, "style": "input", "fmt": MONEY},
        {"name": "Gap (pot.-util.)", "w": 12, "style": "form", "fmt": MONEY},
    ]
    demo = [
        ["M001", "2026-03", 100000, "=C7*ICMS_NOMINAL", "=C7*(PIS_DEBITO+COFINS_DEBITO)", 0, 0, 0, "=D7+E7-G7"],
        ["M002", "2026-03", 50000, "=C8*ICMS_NOMINAL", "=C8*(PIS_DEBITO+COFINS_DEBITO)", 0, 0, 0, "=D8+E8-G8"],
    ]
    table_sheet("11_TAX_CREDITS", "Tax", "Creditos tributarios (potencial->monetizado)",
                "Separar credito potencial, escriturado, utilizado e monetizado. So o efetivamente "
                "recuperavel reduz o Landed Net Cost (aba 09).", cols, demo=demo, gov=True, tab="C55A11")


# ===========================================================================
# 06_BOM  (formulacao tecnica + economica)  &  07_BOM_VERSION
# ===========================================================================
BOM_ROWS = []  # (product, material, row)
def build_bom():
    H = ["PRODUCT_ID", "FORMULA_ID", "Versao", "MATERIAL_ID", "kg / t\nproduto",
         "% massa", "Min\nkg/t", "Max\nkg/t", "Funcao tecnica", "Substituto",
         "LANDED\nR$/t (link)", "Custo padrao\nR$/t produto", "KEY\n(prod|mat)"]
    W = {"A": 12, "B": 10, "C": 7, "D": 12, "E": 9, "F": 8, "G": 7, "H": 7,
         "I": 20, "J": 10, "K": 12, "L": 13, "M": 14}
    ws = custom_head("06_BOM", "Engineering", "Formulacao tecnica e economica (BOM)",
                     "kg/t x LANDED NET R$/t = custo padrao de material. Soma de kg/t deve fechar "
                     "com o peso tecnico (1.000 kg/t). Alerta se nao fechar.", H, W, tab="548235", span=13)
    demo = [
        ("P001", "F-P001", "M001", 250, "Aglomerante"), ("P001", "F-P001", "M002", 700, "Agregado"),
        ("P001", "F-P001", "M003", 45, "Plastificante"), ("P001", "F-P001", "M004", 5, "Retentor agua"),
        ("P002", "F-P002", "M001", 300, "Aglomerante"), ("P002", "F-P002", "M002", 650, "Agregado"),
        ("P002", "F-P002", "M003", 45, "Plastificante"), ("P002", "F-P002", "M004", 5, "Retentor agua"),
        ("P003", "F-P003", "M001", 350, "Aglomerante"), ("P003", "F-P003", "M002", 600, "Agregado"),
        ("P003", "F-P003", "M004", 8, "Retentor agua"), ("P003", "F-P003", "M005", 42, "Pigmento"),
    ]
    lr = "'09_LANDED_COST'!$P${a}:$P${b}".format(a=R0, b=LANDED_LAST)
    la = "'09_LANDED_COST'!$A${a}:$A${b}".format(a=R0, b=LANDED_LAST)
    r = R0
    for prod, fid, mat, kgt, func in demo:
        put(ws, r, 1, prod, style="key", align="c")
        put(ws, r, 2, fid, style="input", align="c")
        put(ws, r, 3, "v1", style="input", align="c")
        put(ws, r, 4, mat, style="key", align="c")
        put(ws, r, 5, kgt, style="input", fmt=NUM, align="r")
        put(ws, r, 6, "=E{r}/1000".format(r=r), style="form", fmt=PCT2, align="r")
        put(ws, r, 7, None, style="input", fmt=NUM)
        put(ws, r, 8, None, style="input", fmt=NUM)
        put(ws, r, 9, func, style="input", align="l")
        put(ws, r, 10, None, style="input", align="c")
        put(ws, r, 11, "=IFERROR(INDEX({lr},MATCH(D{r},{la},0)),0)".format(lr=lr, la=la, r=r),
            style="form", fmt=MONEY, align="r")
        put(ws, r, 12, "=E{r}/1000*K{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 13, "=A{r}&\"|\"&D{r}".format(r=r), style="key", align="c")
        BOM_ROWS.append((prod, mat, r))
        r += 1
    for r in range(R0 + len(demo), R0 + 60):
        put(ws, r, 1, None, style="key"); put(ws, r, 4, None, style="key")
        for c in (2, 3, 9, 10):
            put(ws, r, c, None, style="input")
        put(ws, r, 5, None, style="input", fmt=NUM)
        put(ws, r, 6, "=IF(E{r}=\"\",\"\",E{r}/1000)".format(r=r), style="form", fmt=PCT2)
        for c in (7, 8):
            put(ws, r, c, None, style="input", fmt=NUM)
        put(ws, r, 11, "=IF(D{r}=\"\",\"\",IFERROR(INDEX({lr},MATCH(D{r},{la},0)),0))".format(lr=lr, la=la, r=r), style="form", fmt=MONEY)
        put(ws, r, 12, "=IF(E{r}=\"\",\"\",E{r}/1000*K{r})".format(r=r), style="form", fmt=MONEY)
        put(ws, r, 13, "=IF(A{r}=\"\",\"\",A{r}&\"|\"&D{r})".format(r=r), style="key")
    lastr = R0 + 60
    # control block: per-product weight closure + material standard cost
    cr = lastr + 2
    c = ws.cell(cr, 1, "CONTROLE — FECHAMENTO DA FORMULACAO (peso tecnico)"); c.font = F_SECT
    cr += 1
    header_row(ws, cr, ["PRODUCT_ID", "Soma kg/t", "Peso alvo", "Fecha?",
                        "Custo padrao\nMATERIAL R$/t"], start=1)
    cr += 1
    ar = "$A${a}:$A${b}".format(a=R0, b=lastr)
    er = "$E${a}:$E${b}".format(a=R0, b=lastr)
    lcol = "$L${a}:$L${b}".format(a=R0, b=lastr)
    global BOM_CTRL_START
    BOM_CTRL_START = cr
    for prod in ("P001", "P002", "P003"):
        put(ws, cr, 1, prod, style="key", align="c")
        put(ws, cr, 2, "=SUMIF({ar},A{cr},{er})".format(ar=ar, er=er, cr=cr), style="form", fmt=NUM, align="r")
        put(ws, cr, 3, "=KG_PALETE/ (KG_PALETE/1000)".format(), style="form", fmt=NUM, align="r")  # =1000
        put(ws, cr, 4, '=IF(ABS(B{cr}-1000)<=1,"OK","VERIFICAR")'.format(cr=cr), style="control", align="c")
        put(ws, cr, 5, "=SUMIF({ar},A{cr},{lcol})".format(ar=ar, lcol=lcol, cr=cr), style="form", fmt=MONEY, align="r")
        cr += 1
    ws.freeze_panes = "A7"
    return ws, lastr


def build_bom_version():
    cols = [
        {"name": "FORMULA_ID", "w": 12, "style": "key"},
        {"name": "PRODUCT_ID", "w": 12, "style": "input"},
        {"name": "Versao", "w": 8, "style": "input"},
        {"name": "Motivo alteracao", "w": 26, "style": "input"},
        {"name": "Vigencia ini", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
        {"name": "Vigencia fim", "w": 12, "style": "input", "fmt": "yyyy-mm-dd"},
        {"name": "Aprovador", "w": 14, "style": "input"},
        {"name": "Status", "w": 10, "style": "input"},
    ]
    demo = [
        ["F-P001", "P001", "v1", "Versao base 2026", "2026-01-01", "", "Eng. Tecnico", "Vigente"],
        ["F-P002", "P002", "v1", "Versao base 2026", "2026-01-01", "", "Eng. Tecnico", "Vigente"],
        ["F-P003", "P003", "v1", "Versao base 2026", "2026-01-01", "", "Eng. Tecnico", "Vigente"],
    ]
    table_sheet("07_BOM_VERSION", "Engineering", "Controle de versoes de formulacao",
                "Historico versionado de formulas. Nunca sobrescrever; criar nova versao com vigencia.",
                cols, demo=demo, tab="548235")


BOM_LASTR = R0 + 60

# ===========================================================================
# 17_PRODUCTION_OF  (ordens de fabricacao)
# ===========================================================================
def build_production_of():
    H = ["OF_ID", "PRODUCT_ID", "FORMULA_ID", "Versao", "RESOURCE_ID", "Equipe",
         "Data", "Hora ini", "Hora fim", "Qtd progr.\n(t)", "Qtd produz.\n(t)",
         "Qtd BOA\n(t)", "Sacos", "Paletes", "Perdas\n(t)", "Reproc.\n(t)",
         "Tempo\n(h)", "t/h", "Obs"]
    W = {"A": 10, "B": 11, "C": 10, "D": 7, "E": 11, "F": 10, "G": 11, "H": 8,
         "I": 8, "J": 10, "K": 10, "L": 9, "M": 8, "N": 8, "O": 8, "P": 8, "Q": 8, "R": 8, "S": 16}
    ws = custom_head("17_PRODUCTION_OF", "Production", "Ordens de fabricacao (OF)",
                     "Produtividade medida por hora inicio/fim da OF/lote (NAO por jornada contratual). "
                     "Reproc. separado. t/h = qtd boa / tempo.", H, W, tab="2E75B6", span=19)
    # good tons totals per product used elsewhere: P001=400, P002=250, P003=120
    demo = [
        ["OF-2603-01", "P001", "F-P001", "v1", "RES-MIST1", "Turma A", "2026-03-05", "06:00", "10:00", 400, 402, 400, None, None, 2, 0.0, None, None, ""],
        ["OF-2603-02", "P002", "F-P002", "v1", "RES-MIST1", "Turma A", "2026-03-06", "06:00", "11:00", 250, 251, 250, None, None, 1, 0.0, None, None, ""],
        ["OF-2603-03", "P003", "F-P003", "v1", "RES-MIST2", "Turma B", "2026-03-07", "07:00", "10:00", 120, 121, 120, None, None, 1, 0.5, None, None, ""],
    ]
    r = R0
    for d in demo:
        for i, v in enumerate(d):
            col = i + 1
            if col in (13, 14, 15, 17, 18):  # formula cols
                continue
            st = "key" if col == 1 else "input"
            put(ws, r, col, v, style=st, fmt=(NUM if col in (10, 11, 12, 16) else None),
                align="c" if col in (1, 2, 3, 4, 5, 7, 8, 9) else "l")
        # sacos = qtd boa t *1000 / kg_saco  (lookup product extras)
        put(ws, r, 13, "=IFERROR(L{r}*1000/INDEX('03_PRODUCTS'!$G${a}:$G${b},MATCH(B{r},'03_PRODUCTS'!$A${a}:$A${b},0)),\"\")".format(r=r, a=R0, b=R0+42), style="form", fmt=INT, align="r")
        put(ws, r, 14, "=IFERROR(M{r}/INDEX('03_PRODUCTS'!$H${a}:$H${b},MATCH(B{r},'03_PRODUCTS'!$A${a}:$A${b},0)),\"\")".format(r=r, a=R0, b=R0+42), style="form", fmt=INT, align="r")
        put(ws, r, 15, "=K{r}-L{r}-P{r}".format(r=r), style="form", fmt=NUM, align="r")  # perdas = produz - boa - reproc
        # tempo (h) = fim - ini
        put(ws, r, 17, '=IFERROR((TIMEVALUE(I{r})-TIMEVALUE(H{r}))*24,"")'.format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 18, '=IFERROR(L{r}/Q{r},"")'.format(r=r), style="form", fmt=NUM, align="r")
        r += 1
    for r in range(R0 + len(demo), R0 + 40):
        put(ws, r, 1, None, style="key")
        for col in range(2, 20):
            if col in (13, 14, 15, 17, 18):
                put(ws, r, col, None, style="form")
            else:
                put(ws, r, col, None, style="input")
    return ws


OF_LAST = R0 + 40


# ===========================================================================
# 18_ACTUAL_CONSUMPTION  (consumos reais por OF x material)
# ===========================================================================
def build_actual_consumption():
    H = ["OF_ID", "PRODUCT_ID", "MATERIAL_ID", "Good tons\n(link)",
         "Act kg/t\n(real)", "Act qtd\n(t)", "Act preco\nR$/t (real)", "Act custo\n(R$)",
         "KEY\n(prod|mat)"]
    W = {"A": 12, "B": 11, "C": 12, "D": 10, "E": 10, "F": 10, "G": 11, "H": 12, "I": 14}
    ws = custom_head("18_ACTUAL_CONSUMPTION", "Production", "Consumos reais por OF x material",
                     "Base do Actual Cost. Act qtd = act kg/t /1000 x good tons. "
                     "Act custo = act qtd x act preco.", H, W, tab="2E75B6", span=9)
    # act kg/t and act price DEMO (small deviations from std to gerar variancias)
    demo = [
        ("OF-2603-01", "P001", "M001", 255, 660), ("OF-2603-01", "P001", "M002", 700, 66),
        ("OF-2603-01", "P001", "M003", 45, 485), ("OF-2603-01", "P001", "M004", 5, 17000),
        ("OF-2603-02", "P002", "M001", 300, 660), ("OF-2603-02", "P002", "M002", 655, 66),
        ("OF-2603-02", "P002", "M003", 45, 485), ("OF-2603-02", "P002", "M004", 5, 17000),
        ("OF-2603-03", "P003", "M001", 350, 660), ("OF-2603-03", "P003", "M002", 600, 66),
        ("OF-2603-03", "P003", "M004", 8, 17000), ("OF-2603-03", "P003", "M005", 42, 8610),
    ]
    ofa = "'17_PRODUCTION_OF'!$A${a}:$A${b}".format(a=R0, b=OF_LAST)
    ofl = "'17_PRODUCTION_OF'!$L${a}:$L${b}".format(a=R0, b=OF_LAST)
    r = R0
    for of, prod, mat, actkg, actpr in demo:
        put(ws, r, 1, of, style="key", align="c")
        put(ws, r, 2, prod, style="input", align="c")
        put(ws, r, 3, mat, style="key", align="c")
        put(ws, r, 4, "=IFERROR(INDEX({ofl},MATCH(A{r},{ofa},0)),0)".format(ofl=ofl, ofa=ofa, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 5, actkg, style="input", fmt=NUM, align="r")
        put(ws, r, 6, "=E{r}/1000*D{r}".format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 7, actpr, style="input", fmt=MONEY, align="r")
        put(ws, r, 8, "=F{r}*G{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 9, "=B{r}&\"|\"&C{r}".format(r=r), style="key", align="c")
        r += 1
    for r in range(R0 + len(demo), R0 + 60):
        put(ws, r, 1, None, style="key"); put(ws, r, 3, None, style="key")
        put(ws, r, 2, None, style="input")
        put(ws, r, 4, "=IF(A{r}=\"\",\"\",IFERROR(INDEX({ofl},MATCH(A{r},{ofa},0)),0))".format(ofl=ofl, ofa=ofa, r=r), style="form", fmt=NUM)
        put(ws, r, 5, None, style="input", fmt=NUM)
        put(ws, r, 6, "=IF(E{r}=\"\",\"\",E{r}/1000*D{r})".format(r=r), style="form", fmt=NUM)
        put(ws, r, 7, None, style="input", fmt=MONEY)
        put(ws, r, 8, "=IF(F{r}=\"\",\"\",F{r}*G{r})".format(r=r), style="form", fmt=MONEY)
        put(ws, r, 9, "=IF(B{r}=\"\",\"\",B{r}&\"|\"&C{r})".format(r=r), style="key")
    return ws


AC_LAST = R0 + 60


PRD = ["P001", "P002", "P003"]
PRD_NAME = {"P001": "Multiuso", "P002": "Argamassa AC-III", "P003": "Rejunte"}
CONV_STD = {"P001": 38, "P002": 42, "P003": 55}   # DEMO R$/t (origem: routing)
CONV_ACT = {"P001": 41, "P002": 42, "P003": 58}   # DEMO R$/t

# ranges
OF_A = "'17_PRODUCTION_OF'!$B${a}:$B${b}".format(a=R0, b=OF_LAST)
OF_L = "'17_PRODUCTION_OF'!$L${a}:$L${b}".format(a=R0, b=OF_LAST)
BOM_A = "'06_BOM'!$A${a}:$A${b}".format(a=R0, b=BOM_LASTR)
BOM_D = "'06_BOM'!$D${a}:$D${b}".format(a=R0, b=BOM_LASTR)
BOM_E = "'06_BOM'!$E${a}:$E${b}".format(a=R0, b=BOM_LASTR)
BOM_L = "'06_BOM'!$L${a}:$L${b}".format(a=R0, b=BOM_LASTR)
LAND_A = "'09_LANDED_COST'!$A${a}:$A${b}".format(a=R0, b=LANDED_LAST)
LAND_P = "'09_LANDED_COST'!$P${a}:$P${b}".format(a=R0, b=LANDED_LAST)
AC_B = "'18_ACTUAL_CONSUMPTION'!$B${a}:$B${b}".format(a=R0, b=AC_LAST)
AC_C = "'18_ACTUAL_CONSUMPTION'!$C${a}:$C${b}".format(a=R0, b=AC_LAST)
AC_E = "'18_ACTUAL_CONSUMPTION'!$E${a}:$E${b}".format(a=R0, b=AC_LAST)
AC_G = "'18_ACTUAL_CONSUMPTION'!$G${a}:$G${b}".format(a=R0, b=AC_LAST)
AC_H = "'18_ACTUAL_CONSUMPTION'!$H${a}:$H${b}".format(a=R0, b=AC_LAST)
AC_KEY = "'18_ACTUAL_CONSUMPTION'!$I${a}:$I${b}".format(a=R0, b=AC_LAST)
BOM_KEY = "'06_BOM'!$M${a}:$M${b}".format(a=R0, b=BOM_LASTR)


# ===========================================================================
# 26_STANDARD_COST
# ===========================================================================
def build_standard_cost():
    H = ["PRODUCT_ID", "Descricao", "Good tons\n(link OF)", "Material std\nR$/t",
         "Conversao std\nR$/t", "TOTAL STD\nR$/t", "TOTAL STD\nR$"]
    W = {"A": 12, "B": 24, "C": 11, "D": 12, "E": 12, "F": 12, "G": 14}
    ws = custom_head("26_STANDARD_COST", "Costing", "Custo padrao (standard cost)",
                     "Material std = SUMIF do BOM (kg/t x landed R$/t). Conversao std vem do routing "
                     "(input aqui). Unidade base: R$/tonelada boa produzida.", H, W, tab="7030A0", span=7)
    r = R0
    for p in PRD:
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, PRD_NAME[p], style="link", align="l")
        put(ws, r, 3, "=SUMIF({a},A{r},{l})".format(a=OF_A, l=OF_L, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 4, "=SUMIF({a},A{r},{l})".format(a=BOM_A, l=BOM_L, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 5, CONV_STD[p], style="input", fmt=MONEY, align="r")
        put(ws, r, 6, "=D{r}+E{r}".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
        put(ws, r, 7, "=F{r}*C{r}".format(r=r), style="form", fmt=MONEY, align="r")
        r += 1
    ws.freeze_panes = "A7"
    return ws


# ===========================================================================
# 27_ACTUAL_COST
# ===========================================================================
def build_actual_cost():
    H = ["PRODUCT_ID", "Good tons\n(link)", "Material act\nR$", "Material act\nR$/t",
         "Conversao act\nR$/t", "Conversao act\nR$", "TOTAL ACT\nR$", "TOTAL ACT\nR$/t"]
    W = {"A": 12, "B": 11, "C": 13, "D": 12, "E": 12, "F": 13, "G": 14, "H": 12}
    ws = custom_head("27_ACTUAL_COST", "Costing", "Custo real (actual cost)",
                     "Material real = SUMIFS de 18_ACTUAL_CONSUMPTION. Conversao real (input; origem "
                     "apontamento). TOTAL ACT R$/t e o KPI Material+Conversao por tonelada boa.",
                     H, W, tab="7030A0", span=8)
    r = R0
    for p in PRD:
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, "=SUMIF({a},A{r},{l})".format(a=OF_A, l=OF_L, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 3, "=SUMIF({b},A{r},{h})".format(b=AC_B, h=AC_H, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 4, "=IFERROR(C{r}/B{r},0)".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 5, CONV_ACT[p], style="input", fmt=MONEY, align="r")
        put(ws, r, 6, "=E{r}*B{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, "=C{r}+F{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 8, "=IFERROR(G{r}/B{r},0)".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
        r += 1
    ws.freeze_panes = "A7"
    return ws


# ===========================================================================
# 28_VARIANCES  (motor de variacoes + reconciliacao)
# ===========================================================================
def build_variances():
    H = ["PRODUCT_ID", "MATERIAL_ID", "Good tons", "Std kg/t", "Act kg/t",
         "Std R$/t", "Act R$/t", "Std qtd (t)", "Act qtd (t)", "Std cost R$",
         "Act cost R$", "PPV R$", "Usage R$", "Residual"]
    W = {"A": 11, "B": 11, "C": 9, "D": 8, "E": 8, "F": 9, "G": 9, "H": 9,
         "I": 9, "J": 11, "K": 11, "L": 10, "M": 10, "N": 9}
    ws = custom_head("28_VARIANCES", "Costing", "Motor de variacoes (PPV/Uso/Mix/Yield/Conversao)",
                     "PPV=(Act-Std preco)xAct qtd ; Usage=(Act-Std qtd)xStd preco. "
                     "Std cost + PPV + Usage = Act cost (Residual->0). Mix/Yield e conversao no bloco resumo.",
                     H, W, tab="7030A0", span=14)
    r = R0
    lines = [(p, m) for (p, m, _row) in BOM_ROWS]
    for p, m in lines:
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, m, style="key", align="c")
        put(ws, r, 3, "=SUMIF({a},A{r},{l})".format(a=OF_A, l=OF_L, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 4, "=IFERROR(INDEX({e},MATCH(A{r}&\"|\"&B{r},{k},0)),0)".format(k=BOM_KEY, e=BOM_E, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 5, "=IFERROR(INDEX({e},MATCH(A{r}&\"|\"&B{r},{k},0)),0)".format(k=AC_KEY, e=AC_E, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 6, "=IFERROR(INDEX({p},MATCH(B{r},{a},0)),0)".format(p=LAND_P, a=LAND_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, "=IFERROR(INDEX({g},MATCH(A{r}&\"|\"&B{r},{k},0)),0)".format(g=AC_G, k=AC_KEY, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 8, "=D{r}/1000*C{r}".format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 9, "=E{r}/1000*C{r}".format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 10, "=H{r}*F{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 11, "=I{r}*G{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 12, "=(G{r}-F{r})*I{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 13, "=(I{r}-H{r})*F{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 14, "=K{r}-J{r}-L{r}-M{r}".format(r=r), style="control", fmt=MONEY4, align="r")
        r += 1
    vend = r - 1
    global VAR_A, VAR_LAST
    VAR_A = "$A${a}:$A${b}".format(a=R0, b=vend)
    VAR_LAST = vend
    # product-level summary: Mix/Yield split + conversion + reconciliation
    sr = r + 2
    c = ws.cell(sr, 1, "RESUMO POR PRODUTO — MIX/YIELD, CONVERSAO E RECONCILIACAO"); c.font = F_SECT
    sr += 1
    SH = ["PRODUCT_ID", "Std mat R$", "Act mat R$", "PPV R$", "Usage R$",
          "Std qtd T", "Act qtd T", "wsp R$/t", "Yield R$", "Mix R$",
          "Conv var R$", "Total var R$", "Recon check"]
    header_row(ws, sr, SH, start=1)
    sr += 1
    global VAR_SUM_START
    VAR_SUM_START = sr
    aJ = "$J${a}:$J${b}".format(a=R0, b=vend)
    aK = "$K${a}:$K${b}".format(a=R0, b=vend)
    aL = "$L${a}:$L${b}".format(a=R0, b=vend)
    aM = "$M${a}:$M${b}".format(a=R0, b=vend)
    aH = "$H${a}:$H${b}".format(a=R0, b=vend)
    aI = "$I${a}:$I${b}".format(a=R0, b=vend)
    for p in PRD:
        put(ws, sr, 1, p, style="key", align="c")
        put(ws, sr, 2, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aJ, r=sr), style="form", fmt=MONEY, align="r")
        put(ws, sr, 3, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aK, r=sr), style="form", fmt=MONEY, align="r")
        put(ws, sr, 4, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aL, r=sr), style="form", fmt=MONEY, align="r")
        put(ws, sr, 5, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aM, r=sr), style="form", fmt=MONEY, align="r")
        put(ws, sr, 6, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aH, r=sr), style="form", fmt=NUM, align="r")
        put(ws, sr, 7, "=SUMIF({a},A{r},{v})".format(a=VAR_A, v=aI, r=sr), style="form", fmt=NUM, align="r")
        put(ws, sr, 8, "=IFERROR(B{r}/F{r},0)".format(r=sr), style="form", fmt=MONEY, align="r")  # wsp=StdMat/StdQty
        put(ws, sr, 9, "=(G{r}-F{r})*H{r}".format(r=sr), style="form", fmt=MONEY, align="r")  # Yield=(Ta-Ts)*wsp
        put(ws, sr, 10, "=E{r}-I{r}".format(r=sr), style="form", fmt=MONEY, align="r")  # Mix=Usage-Yield
        put(ws, sr, 11, "=(INDEX('27_ACTUAL_COST'!$E${a}:$E${b},MATCH(A{r},'27_ACTUAL_COST'!$A${a}:$A${b},0))-INDEX('26_STANDARD_COST'!$E${a}:$E${b},MATCH(A{r},'26_STANDARD_COST'!$A${a}:$A${b},0)))*INDEX('26_STANDARD_COST'!$C${a}:$C${b},MATCH(A{r},'26_STANDARD_COST'!$A${a}:$A${b},0))".format(a=R0, b=R0+len(PRD)-1, r=sr), style="form", fmt=MONEY, align="r")
        put(ws, sr, 12, "=D{r}+E{r}+K{r}".format(r=sr), style="form", fmt=MONEY, align="r")  # PPV+Usage+ConvVar
        # recon: Std total (26!G) + Total var (L) - Act total (27!G)  ->  0
        put(ws, sr, 13, "=ROUND(INDEX('26_STANDARD_COST'!$G${a}:$G${b},MATCH(A{r},'26_STANDARD_COST'!$A${a}:$A${b},0))+L{r}-INDEX('27_ACTUAL_COST'!$G${a}:$G${b},MATCH(A{r},'27_ACTUAL_COST'!$A${a}:$A${b},0)),2)".format(a=R0, b=R0+len(PRD)-1, r=sr), style="control", fmt=MONEY4, align="r")
        sr += 1
    # recon check simpler note
    put(ws, sr + 1, 1, "Recon check = (Std total + Total var) - Act total  ->  deve ser 0,00 por construcao.",
        style="note", border=False)
    ws.freeze_panes = "A7"
    return ws


def C(name, w=14, style="input", fmt=None):
    d = {"name": name, "w": w, "style": style}
    if fmt:
        d["fmt"] = fmt
    return d


PROC = "C55A11"; PROD = "2E75B6"; COST = "7030A0"; COMM = "548235"; FIN = "BF8F00"


def build_structured_batch():
    specs = [
        # 08_PURCHASES
        ("08_PURCHASES", "Procurement", "Compras reais (notas)",
         "Base bruta de compras. Alimenta 09_LANDED_COST e PPV (preco real vs padrao).",
         [C("PURCHASE_ID", 12, "key"), C("MATERIAL_ID", 12), C("SUPPLIER_ID", 12),
          C("Data", 12, fmt="yyyy-mm-dd"), C("NF", 10), C("Qtd", 10, fmt=NUM),
          C("Unid", 6), C("Preco bruto R$/un", 12, fmt=MONEY4), C("Desconto R$", 10, fmt=MONEY),
          C("Frete R$", 10, fmt=MONEY), C("ICMS destac. R$", 12, fmt=MONEY),
          C("PIS/COFINS R$", 12, fmt=MONEY)],
         [["PC-001", "M001", "S001", "2026-03-02", "12345", 30, "t", 700, 600, 1350, 3780, 0],
          ["PC-002", "M002", "S002", "2026-03-03", "12346", 60, "t", 55, 0, 1320, 594, 0]], True, PROC),
        # 12_RESOURCES
        ("12_RESOURCES", "Resources", "Maquinas/equipamentos",
         "Base do Resource Cost Model e da capacidade pratica.",
         [C("RESOURCE_ID", 12, "key"), C("Descricao", 24), C("COST_CENTER_ID", 12),
          C("Cap. nominal (t/h)", 12, fmt=NUM), C("Cap. pratica (t/h)", 12, fmt=NUM),
          C("Horas disp./mes", 12, fmt=NUM), C("Depreciacao R$/mes", 14, fmt=MONEY),
          C("Manut. fixa R$/mes", 14, fmt=MONEY), C("Outros fixos R$/mes", 14, fmt=MONEY)],
         [["RES-MIST1", "Misturador industrial 1", "CC-PROD", 8.0, 6.5, 360, 8000, 3000, 2000],
          ["RES-MIST2", "Misturador industrial 2", "CC-PROD", 6.0, 5.0, 360, 6000, 2500, 1500]], True, COST),
        # 13_COST_CENTERS
        ("13_COST_CENTERS", "Resources", "Centros de custo",
         "Estrutura de acumulacao. Reprocesso tem centro proprio (nao contamina equipe principal).",
         [C("COST_CENTER_ID", 12, "key"), C("Descricao", 24), C("Tipo", 14),
          C("Responsavel", 16), C("Custo fixo mes R$", 14, fmt=MONEY)],
         [["CC-PROD", "Producao principal", "Produtivo", "Sup. Producao", 30000],
          ["CC-REPRO", "Reprocesso", "Produtivo (separado)", "Sup. Producao", 4000],
          ["CC-LAB", "Laboratorio/Qualidade", "Apoio", "Qualidade", 6000],
          ["CC-MANUT", "Manutencao", "Apoio", "Manutencao", 8000]], True, COST),
        # 14_ROUTINGS
        ("14_ROUTINGS", "Resources", "Roteiros produtivos",
         "Etapas: dosagem, mistura, transferencia, ensacamento, paletizacao, lab, movim., armazenag. "
         "Alimenta Time-Driven ABC (taxa x tempo consumido).",
         [C("ROUTING_ID", 10, "key"), C("PRODUCT_ID", 11), C("Seq", 6, fmt=INT),
          C("Etapa", 16), C("RESOURCE_ID", 12), C("Tempo padrao (h/t)", 12, fmt=MONEY4),
          C("Cap (t/h)", 10, fmt=NUM)],
         [["RT-EXT", "P001", 1, "Dosagem", "RES-MIST1", 0.06, 8],
          ["RT-EXT", "P001", 2, "Mistura", "RES-MIST1", 0.10, 6.5],
          ["RT-EXT", "P001", 3, "Ensacamento", "RES-MIST1", 0.05, 8],
          ["RT-ARG", "P002", 1, "Mistura", "RES-MIST1", 0.12, 6.5],
          ["RT-REJ", "P003", 1, "Mistura", "RES-MIST2", 0.16, 5]], False, COST),
        # 19_LOSSES
        ("19_LOSSES", "Production", "Perdas (normal/anormal)",
         "Mass balance: entrada = bom + perda normal + perda anormal + reproc + WIP + nao explicado. "
         "Diferenca nao explicada (UNACCOUNTED) nunca e absorvida automaticamente.",
         [C("OF_ID", 12, "key"), C("PRODUCT_ID", 11), C("Entrada MP (t)", 12, fmt=NUM),
          C("Bom (t)", 10, fmt=NUM), C("Perda normal (t)", 12, fmt=NUM),
          C("Perda anormal (t)", 12, fmt=NUM), C("Reproc (t)", 10, fmt=NUM),
          C("WIP (t)", 8, fmt=NUM), C("Nao explicado (t)", 12, "form", fmt=NUM)],
         [["OF-2603-01", "P001", 402, 400, 2, 0, 0, 0, "=C7-D7-E7-F7-G7-H7"],
          ["OF-2603-02", "P002", 251, 250, 1, 0, 0, 0, "=C8-D8-E8-F8-G8-H8"],
          ["OF-2603-03", "P003", 121.5, 120, 1, 0, 0.5, 0, "=C9-D9-E9-F9-G9-H9"]], False, PROD),
        # 20_REPROCESS
        ("20_REPROCESS", "Production", "Reprocessamentos (custo economico separado)",
         "Excluir reproc dos indicadores da equipe principal; NAO excluir seu custo economico. "
         "Net Reprocess = custo perdido + recuperacao - valor recuperado.",
         [C("REPRO_ID", 10, "key"), C("OF_ID origem", 12), C("PRODUCT_ID", 11),
          C("Qtd (t)", 8, fmt=NUM), C("Custo original perdido R$", 16, fmt=MONEY),
          C("Custo recuperacao R$", 14, fmt=MONEY), C("Valor recuperado R$", 14, fmt=MONEY),
          C("NET reprocess R$", 14, "form", fmt=MONEY)],
         [["RP-001", "OF-2603-03", "P003", 0.5, 150, 60, 120, "=E7+F7-G7"]], False, PROD),
        # 21_ENERGY
        ("21_ENERGY", "Production", "Energia (tarifa/demanda/consumo)",
         "Energy Price Variance e Energy Usage Variance. kWh/t por OF quando disponivel.",
         [C("PERIOD", 10, "key"), C("RESOURCE_ID", 12), C("kWh consumidos", 12, fmt=NUM),
          C("Tarifa R$/kWh", 12, fmt=MONEY4), C("Demanda R$", 10, fmt=MONEY),
          C("Producao t", 10, fmt=NUM), C("kWh/t", 8, "form", fmt=NUM),
          C("Custo energia R$", 12, "form", fmt=MONEY)],
         [["2026-03", "RES-MIST1", 9000, 0.72, 1500, 650, "=IFERROR(C7/F7,0)", "=C7*D7+E7"]], False, PROD),
        # 22_LABOR
        ("22_LABOR", "Production", "Mao de obra (estrutural/variavel)",
         "Separar hora normal, extra, setup, espera, reprocesso, absenteismo. Labor Efficiency Variance.",
         [C("PERIOD", 10, "key"), C("COST_CENTER_ID", 12), C("Horas normais", 12, fmt=NUM),
          C("Horas extras", 12, fmt=NUM), C("Taxa normal R$/h", 12, fmt=MONEY),
          C("Taxa extra R$/h", 12, fmt=MONEY), C("Horas padrao", 12, fmt=NUM),
          C("Custo MO R$", 12, "form", fmt=MONEY)],
         [["2026-03", "CC-PROD", 480, 40, 22, 33, 470, "=C7*E7+D7*F7"]], False, PROD),
        # 23_MAINTENANCE
        ("23_MAINTENANCE", "Production", "Manutencao (prev/pred/corr/emerg)",
         "Maintenance Cost/Running Hour e Breakdown Economic Cost (inclui perda de producao).",
         [C("PERIOD", 10, "key"), C("RESOURCE_ID", 12), C("Tipo", 12), C("Pecas R$", 10, fmt=MONEY),
          C("Servicos R$", 10, fmt=MONEY), C("Horas parada", 10, fmt=NUM),
          C("Perda producao R$", 12, fmt=MONEY), C("Custo total R$", 12, "form", fmt=MONEY)],
         [["2026-03", "RES-MIST1", "Corretiva", 800, 400, 3, 900, "=D7+E7+G7"]], False, PROD),
        # 24_QUALITY
        ("24_QUALITY", "Quality", "Qualidade (4 classes COPQ)",
         "Prevention, Appraisal, Internal Failure, External Failure. Alimenta 11_COPQ / dashboards.",
         [C("PERIOD", 10, "key"), C("PRODUCT_ID", 11), C("Classe", 16), C("Categoria", 18),
          C("Valor R$", 12, fmt=MONEY), C("Causa", 16), C("Responsavel area", 14)],
         [["2026-03", "P003", "Internal Failure", "Reprocesso", 90, "Ajuste cor", "Producao"],
          ["2026-03", "P001", "Appraisal", "Ensaio lab", 200, "Controle rotina", "Qualidade"]], False, PROD),
        # 25_INVENTORY
        ("25_INVENTORY", "Inventory", "Estoques (MP/WIP/PA) e carrying cost",
         "Dias, giro, estoque medio + Inventory Carrying Cost (capital + armazenagem + seguro + obsoles.).",
         [C("ITEM_ID", 12, "key"), C("Tipo", 10), C("Estoque medio R$", 14, fmt=MONEY),
          C("Dias", 8, fmt=NUM), C("Giro", 8, fmt=NUM), C("Armazenagem %", 10, fmt=PCT2),
          C("Carrying cost R$", 14, "form", fmt=MONEY)],
         [["M001", "MP", 64000, 30, 12, 0.03, "=C7*(CUSTO_CAPITAL_ANO+F7)*D7/365"],
          ["P001", "PA", 30000, 15, 24, 0.03, "=C8*(CUSTO_CAPITAL_ANO+F8)*D8/365"]], False, "203864"),
        # 30_CONSTRAINT
        ("30_CONSTRAINT", "Capacity", "Gargalo (TOC) e MC/hora",
         "Identifica o CONSTRAINT RESOURCE e a Margem de Contribuicao/hora do gargalo (mix, pedidos especiais).",
         [C("RESOURCE_ID", 12, "key"), C("E gargalo?", 10), C("PRODUCT_ID", 11),
          C("MC unit R$/t", 12, fmt=MONEY), C("Tempo gargalo h/t", 12, fmt=MONEY4),
          C("MC / hora gargalo R$", 16, "form", fmt=MONEY)],
         [["RES-MIST1", "Sim", "P001", 120, 0.21, "=IFERROR(D7/E7,0)"],
          ["RES-MIST1", "Sim", "P002", 140, 0.12, "=IFERROR(D8/E8,0)"]], False, "203864"),
        # 33_SALES
        ("33_SALES", "Commercial", "Vendas (cliente x SKU)",
         "Unidade minima de rentabilidade: CLIENTE x SKU.",
         [C("SALE_ID", 10, "key"), C("CUSTOMER_ID", 12), C("PRODUCT_ID", 11),
          C("Data", 11, fmt="yyyy-mm-dd"), C("Qtd t", 8, fmt=NUM), C("Preco tabela R$/t", 12, fmt=MONEY),
          C("Desconto R$/t", 10, fmt=MONEY), C("Bonific R$/t", 10, fmt=MONEY),
          C("Comissao %", 8, fmt=PCT2), C("Prazo dias", 8, fmt=INT)],
         [["V-001", "C001", "P001", "2026-03-10", 20, 520, 10, 5, 0.02, 30],
          ["V-002", "C002", "P002", "2026-03-11", 15, 560, 20, 0, 0.02, 45],
          ["V-003", "C003", "P003", "2026-03-12", 5, 900, 0, 0, 0.03, 30]], False, COMM),
        # 34_FREIGHT_OUT
        ("34_FREIGHT_OUT", "Commercial", "Frete de saida",
         "Frete absorvido entra no Cost to Serve e no Price Waterfall.",
         [C("SALE_ID", 10, "key"), C("CUSTOMER_ID", 12), C("Frete total R$", 12, fmt=MONEY),
          C("Absorvido pela SKAL?", 14), C("R$/t", 10, "form", fmt=MONEY)],
         [["V-001", "C001", 800, "Sim", "=IFERROR(C7/INDEX('33_SALES'!$E$7:$E$40,MATCH(A7,'33_SALES'!$A$7:$A$40,0)),0)"]], False, COMM),
        # 35_COMMISSIONS
        ("35_COMMISSIONS", "Commercial", "Comissoes",
         "Comissao por venda; entra no Cost to Serve.",
         [C("SALE_ID", 10, "key"), C("Vendedor", 14), C("% comissao", 10, fmt=PCT2),
          C("Base R$", 12, fmt=MONEY), C("Comissao R$", 12, "form", fmt=MONEY)],
         [["V-001", "Rep 1", 0.02, 7200, "=C7*D7"]], False, COMM),
        # 36_RECEIVABLES
        ("36_RECEIVABLES", "Commercial", "Prazo e inadimplencia",
         "Receivable Financing Cost = valor x taxa x dias/365. Expected Credit Loss = PD x exposicao x LGD.",
         [C("SALE_ID", 10, "key"), C("Valor a receber R$", 14, fmt=MONEY), C("Dias", 8, fmt=INT),
          C("Custo prazo R$", 12, "form", fmt=MONEY), C("PD %", 8, fmt=PCT2), C("LGD %", 8, fmt=PCT2),
          C("Expected credit loss R$", 16, "form", fmt=MONEY)],
         [["V-001", 7000, 30, "=B7*CUSTO_CAPITAL_ANO*C7/DIAS_ANO", 0.015, 0.6, "=B7*E7*F7"]], False, COMM),
        # 37_BARTER
        ("37_BARTER", "Commercial", "Permutas (valor economico liquido)",
         "Nao tratar permuta pelo valor nominal. Liquido = valor econom. - monetizacao - custo fin. - risco - trib.",
         [C("BARTER_ID", 10, "key"), C("Contraparte", 16), C("Valor nominal R$", 12, fmt=MONEY),
          C("Valor economico R$", 14, fmt=MONEY), C("Custo monetizacao R$", 14, fmt=MONEY),
          C("Custo financeiro R$", 14, fmt=MONEY), C("Risco R$", 10, fmt=MONEY),
          C("Trib. adic R$", 10, fmt=MONEY), C("Liquido R$", 12, "form", fmt=MONEY)],
         None, False, COMM),
        # 41_RECON_PHYSICAL
        ("41_RECON_PHYSICAL", "Reconciliation", "Reconciliacao fisica de estoque",
         "Est. inicial + compras - consumo - perdas +/- transf = est. final esperado; comparar com real.",
         [C("MATERIAL_ID", 12, "key"), C("Est. inicial t", 12, fmt=NUM), C("Compras t", 10, fmt=NUM),
          C("Consumo t", 10, fmt=NUM), C("Perdas t", 10, fmt=NUM), C("Transf t", 10, fmt=NUM),
          C("Est. final esperado t", 14, "form", fmt=NUM), C("Est. final real t", 12, fmt=NUM),
          C("Diferenca t", 10, "form", fmt=NUM)],
         [["M001", 100, 30, 90.15, 0, 0, "=B7+C7-D7-E7+F7", 39.85, "=H7-G7"]], False, "203864"),
    ]
    for name, grp, purp, sub, cols, demo, gov, tab in specs:
        table_sheet(name, grp, purp, sub, cols, demo=demo, gov=gov, tab=tab)


TB = R0 + 60  # generic table data end used for lookups (table_sheet nrows=60)


def build_resource_rates():
    H = ["RESOURCE_ID", "Custo fixo R$/mes\n(dep+manut+outros)", "Horas praticas/mes",
         "R$/hora pratica"]
    W = {"A": 12, "B": 20, "C": 16, "D": 14}
    ws = custom_head("15_RESOURCE_RATES", "Resources", "R$/hora de capacidade pratica",
                     "R$/hora = custo fixo / horas de CAPACIDADE PRATICA (nao a producao real). "
                     "Base do custo de conversao e do unused capacity cost.", H, W, tab=COST, span=4)
    r12a = "'12_RESOURCES'!$A${a}:$A${b}".format(a=R0, b=TB)
    for i, res in enumerate(["RES-MIST1", "RES-MIST2"]):
        r = R0 + i
        put(ws, r, 1, res, style="key", align="c")
        put(ws, r, 2, "=SUMIF({k},A{r},'12_RESOURCES'!$G${a}:$G${b})+SUMIF({k},A{r},'12_RESOURCES'!$H${a}:$H${b})+SUMIF({k},A{r},'12_RESOURCES'!$I${a}:$I${b})".format(a=R0, b=TB, k=r12a, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 3, "=SUMIF({k},A{r},'12_RESOURCES'!$F${a}:$F${b})".format(a=R0, b=TB, k=r12a, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 4, "=IFERROR(B{r}/C{r},0)".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
    return ws


def build_standard_times():
    H = ["PRODUCT_ID", "Tempo padrao total h/t\n(soma routing)", "Velocidade padrao t/h"]
    W = {"A": 12, "B": 20, "C": 18}
    ws = custom_head("16_STANDARD_TIMES", "Resources", "Tempos padrao por produto",
                     "Consolida o routing (14) em h/t por produto. Base do custo de conversao padrao "
                     "e do speed loss.", H, W, tab=COST, span=3)
    r14b = "'14_ROUTINGS'!$B${a}:$B${b}".format(a=R0, b=TB)
    r14f = "'14_ROUTINGS'!$F${a}:$F${b}".format(a=R0, b=TB)
    for i, p in enumerate(PRD):
        r = R0 + i
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, "=SUMIF({b},A{r},{f})".format(b=r14b, f=r14f, r=r), style="form", fmt=MONEY4, align="r")
        put(ws, r, 3, "=IFERROR(1/B{r},0)".format(r=r), style="form", fmt=NUM, align="r")
    return ws


def build_capacity():
    H = ["RESOURCE_ID", "Horas praticas\n(link)", "Horas usadas\n(link OF)",
         "Horas ociosas", "R$/hora\n(link)", "UNUSED CAPACITY\nCOST R$"]
    W = {"A": 12, "B": 12, "C": 12, "D": 12, "E": 10, "F": 16}
    ws = custom_head("29_CAPACITY", "Capacity", "Capacidade e ociosidade (unused capacity)",
                     "Horas ociosas = praticas - usadas. Unused Capacity Cost = ociosas x R$/hora. "
                     "Este custo fica VISIVEL, nao e rateado ao produto.", H, W, tab="203864", span=6)
    r12a = "'12_RESOURCES'!$A${a}:$A${b}".format(a=R0, b=TB)
    r15a = "'15_RESOURCE_RATES'!$A${a}:$A${b}".format(a=R0, b=R0 + 2)
    ofE = "'17_PRODUCTION_OF'!$E${a}:$E${b}".format(a=R0, b=OF_LAST)
    ofQ = "'17_PRODUCTION_OF'!$Q${a}:$Q${b}".format(a=R0, b=OF_LAST)
    for i, res in enumerate(["RES-MIST1", "RES-MIST2"]):
        r = R0 + i
        put(ws, r, 1, res, style="key", align="c")
        put(ws, r, 2, "=SUMIF({k},A{r},'12_RESOURCES'!$F${a}:$F${b})".format(a=R0, b=TB, k=r12a, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 3, "=SUMIF({e},A{r},{q})".format(q=ofQ, e=ofE, r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 4, "=B{r}-C{r}".format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 5, "=IFERROR(INDEX('15_RESOURCE_RATES'!$D${a}:$D${b},MATCH(A{r},{k},0)),0)".format(a=R0, b=R0 + 2, k=r15a, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 6, "=D{r}*E{r}".format(r=r), style="alert", fmt=MONEY, align="r", bold=True)
    put(ws, R0 + 3, 1, "Total unused capacity cost:", style="note", border=False)
    put(ws, R0 + 3, 6, "=SUM(F{a}:F{b})".format(a=R0, b=R0 + 1), style="control", fmt=MONEY, align="r", bold=True)
    return ws


STD_F = "'26_STANDARD_COST'!$F${a}:$F${b}".format(a=R0, b=R0 + 2)
STD_A = "'26_STANDARD_COST'!$A${a}:$A${b}".format(a=R0, b=R0 + 2)
ACT_H = "'27_ACTUAL_COST'!$H${a}:$H${b}".format(a=R0, b=R0 + 2)
ACT_A = "'27_ACTUAL_COST'!$A${a}:$A${b}".format(a=R0, b=R0 + 2)


# ===========================================================================
# 31_COST_GAP  (5 niveis de custo + gaps)
# ===========================================================================
def build_cost_gap():
    H = ["PRODUCT_ID", "1.Actual\nR$/t", "2.Frozen Std\nR$/t", "3.Current Std\nR$/t",
         "4.Replacement\nR$/t", "5.Best Demonstr.\nR$/t", "6.Entitlement\nR$/t",
         "7.Theoretical\nR$/t", "Gap Act-\nCurStd", "Gap Act-\nBest", "Gap Act-\nEntitle",
         "Vol t/ano", "CONTROLAVEL\nR$/ano"]
    W = {"A": 11}
    for col in "BCDEFGHIJKLM":
        W[col] = 11
    ws = custom_head("31_COST_GAP", "Costing", "Cost gap — 5 niveis de custo",
                     "Actual -> Current Std -> Best Demonstrated -> Entitlement -> Theoretical. "
                     "Controlavel/ano = (Actual - Entitlement) x volume anual. Amarelo = input (a preencher).",
                     H, W, tab=COST, span=13)
    # DEMO frozen/replacement/best/entitlement/theoretical (a preencher com dados reais)
    demo = {"P001": (None, None, None, None, None), "P002": (None, None, None, None, None),
            "P003": (None, None, None, None, None)}
    for i, p in enumerate(PRD):
        r = R0 + i
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, "=IFERROR(INDEX({h},MATCH(A{r},{a},0)),0)".format(h=ACT_H, a=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 3, None, style="input", fmt=MONEY, align="r")  # frozen std
        put(ws, r, 4, "=IFERROR(INDEX({f},MATCH(A{r},{a},0)),0)".format(f=STD_F, a=STD_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 5, None, style="input", fmt=MONEY, align="r")  # replacement
        put(ws, r, 6, None, style="input", fmt=MONEY, align="r")  # best demonstrated
        put(ws, r, 7, None, style="input", fmt=MONEY, align="r")  # entitlement
        put(ws, r, 8, None, style="input", fmt=MONEY, align="r")  # theoretical
        put(ws, r, 9, "=B{r}-D{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 10, '=IF(F{r}="","",B{r}-F{r})'.format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 11, '=IF(G{r}="","",B{r}-G{r})'.format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 12, None, style="input", fmt=INT, align="r")  # volume ano
        put(ws, r, 13, '=IF(OR(G{r}="",L{r}=""),"",(B{r}-G{r})*L{r})'.format(r=r), style="control", fmt=INT, align="r")
        r += 1
    put(ws, R0 + 4, 1, "Preencha Frozen Std, Replacement, Best Demonstrated, Entitlement, Theoretical e Volume "
        "(ver DATA_GAPS). Actual e Current Std sao calculados.", style="note", border=False)
    return ws


# ===========================================================================
# 38_COST_TO_SERVE
# ===========================================================================
def build_cost_to_serve():
    H = ["SALE_ID", "PRODUCT_ID", "Qtd t", "Separacao R$", "Carregamento R$",
         "Frete R$\n(link)", "Comissao R$\n(link)", "Custo prazo R$\n(link)",
         "Inadimpl R$\n(link)", "Bonific/Devol R$", "COST TO SERVE R$", "R$/t"]
    W = {"A": 10, "B": 11, "C": 8, "D": 10, "E": 12, "F": 10, "G": 11, "H": 12, "I": 11, "J": 12, "K": 13, "L": 9}
    ws = custom_head("38_COST_TO_SERVE", "Commercial", "Custo de servir (cost to serve)",
                     "Apos o custo industrial, agrega custos de servir por venda. Nota: no Pocket Price (40) "
                     "os itens comerciais ja sao deduzidos; aqui e a visao consolidada do cost to serve.",
                     H, W, tab=COMM, span=12)
    s34 = ("'34_FREIGHT_OUT'!$A${a}:$A${b}".format(a=R0, b=TB), "'34_FREIGHT_OUT'!$C${a}:$C${b}".format(a=R0, b=TB))
    s35 = ("'35_COMMISSIONS'!$A${a}:$A${b}".format(a=R0, b=TB), "'35_COMMISSIONS'!$E${a}:$E${b}".format(a=R0, b=TB))
    s36c = ("'36_RECEIVABLES'!$A${a}:$A${b}".format(a=R0, b=TB), "'36_RECEIVABLES'!$D${a}:$D${b}".format(a=R0, b=TB))
    s36i = ("'36_RECEIVABLES'!$A${a}:$A${b}".format(a=R0, b=TB), "'36_RECEIVABLES'!$G${a}:$G${b}".format(a=R0, b=TB))
    sales = [("V-001", "P001", 20), ("V-002", "P002", 15), ("V-003", "P003", 5)]
    for i, (sid, p, q) in enumerate(sales):
        r = R0 + i
        put(ws, r, 1, sid, style="key", align="c")
        put(ws, r, 2, p, style="link", align="c")
        put(ws, r, 3, q, style="link", fmt=NUM, align="r")
        put(ws, r, 4, 40 if i == 0 else 30, style="input", fmt=MONEY, align="r")
        put(ws, r, 5, 60 if i == 0 else 40, style="input", fmt=MONEY, align="r")
        put(ws, r, 6, "=IFERROR(SUMIF({a},A{r},{v}),0)".format(a=s34[0], v=s34[1], r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, "=IFERROR(SUMIF({a},A{r},{v}),0)".format(a=s35[0], v=s35[1], r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 8, "=IFERROR(SUMIF({a},A{r},{v}),0)".format(a=s36c[0], v=s36c[1], r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 9, "=IFERROR(SUMIF({a},A{r},{v}),0)".format(a=s36i[0], v=s36i[1], r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 10, 0, style="input", fmt=MONEY, align="r")
        put(ws, r, 11, "=SUM(D{r}:J{r})".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
        put(ws, r, 12, "=IFERROR(K{r}/C{r},0)".format(r=r), style="form", fmt=MONEY, align="r")
    return ws


# ===========================================================================
# 40_PROFITABILITY  (cliente x SKU, pocket price/margin)
# ===========================================================================
def build_profitability():
    H = ["SALE_ID", "CUST", "PROD", "Qtd t", "Preco tabela\nR$/t", "(-)Desc\nR$/t",
         "(-)Bonif\nR$/t", "(-)Comissao\nR$/t", "(-)Frete\nR$/t", "(-)Custo prazo\nR$/t",
         "(-)Inadimpl\nR$/t", "POCKET PRICE\nR$/t", "(-)Custo indl\nR$/t", "(-)C.serve extra\nR$/t",
         "(-)Trib econ\nR$/t", "POCKET MARGIN\nR$/t", "Margem %", "Total R$"]
    W = {"A": 9, "B": 6, "C": 6}
    for col in "DEFGHIJKLMNOPQR":
        W[col] = 10
    ws = custom_head("40_PROFITABILITY", "Commercial", "Rentabilidade cliente x SKU (pocket margin)",
                     "Price waterfall -> Pocket Price -> Pocket Margin. Custo industrial = custo real R$/t (27). "
                     "Trib economica = pocket x ICMS efetivo. Verde = margem>0, vermelho = destruidor de valor.",
                     H, W, tab=COMM, span=18)
    sa = "'33_SALES'!$A${a}:$A${b}".format(a=R0, b=TB)
    def idx(col):
        return "INDEX('33_SALES'!${c}${a}:${c}${b},MATCH(A{{r}},{sa},0))".format(c=col, a=R0, b=TB, sa=sa)
    for i, sid in enumerate(["V-001", "V-002", "V-003"]):
        r = R0 + i
        put(ws, r, 1, sid, style="key", align="c")
        put(ws, r, 2, "=" + idx("B").format(r=r), style="form", align="c")
        put(ws, r, 3, "=" + idx("C").format(r=r), style="form", align="c")
        put(ws, r, 4, "=" + idx("E").format(r=r), style="form", fmt=NUM, align="r")
        put(ws, r, 5, "=" + idx("F").format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 6, "=" + idx("G").format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, "=" + idx("H").format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 8, "=E{r}*{ci}".format(r=r, ci=idx("I").format(r=r)), style="form", fmt=MONEY, align="r")  # comissao R$/t
        put(ws, r, 9, 40 if i == 0 else 0, style="input", fmt=MONEY, align="r")  # frete R$/t absorvido
        put(ws, r, 10, "=(E{r}-F{r}-G{r})*CUSTO_CAPITAL_ANO*{dz}/DIAS_ANO".format(r=r, dz=idx("J").format(r=r)), style="form", fmt=MONEY, align="r")
        put(ws, r, 11, "=(E{r}-F{r}-G{r})*INADIMPL_PCT".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 12, "=E{r}-F{r}-G{r}-H{r}-I{r}-J{r}-K{r}".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
        put(ws, r, 13, "=IFERROR(INDEX({h},MATCH(C{r},{a},0)),0)".format(h=ACT_H, a=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 14, 5, style="input", fmt=MONEY, align="r")  # cost to serve extra (separacao/carreg) R$/t
        put(ws, r, 15, "=L{r}*ICMS_EFETIVO".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 16, "=L{r}-M{r}-N{r}-O{r}".format(r=r), style="form", fmt=MONEY, align="r", bold=True)
        put(ws, r, 17, "=IFERROR(P{r}/E{r},0)".format(r=r), style="form", fmt=PCT2, align="r")
        put(ws, r, 18, "=P{r}*D{r}".format(r=r), style="form", fmt=MONEY, align="r")
        # conditional visual: mark destroyer
        put(ws, r, 19, '=IF(P{r}<0,"DESTRUIDOR DE VALOR",IF(Q{r}<0.05,"MARGEM BAIXA","OK"))'.format(r=r), style="control", align="c")
    ws.cell(6, 19, "Flag"); ws.cell(6, 19).font = F_HEADER; ws.cell(6, 19).fill = FILL_HEADER
    ws.cell(6, 19).alignment = AL_C; ws.cell(6, 19).border = BORDER_ALL
    W2 = {"S": 18}
    widths(ws, W2)
    return ws


# ===========================================================================
# 39_PRICING  (7 niveis de preco, definicoes)
# ===========================================================================
def build_pricing():
    H = ["PRODUCT_ID", "Custo variavel\nR$/t", "Custo total\nR$/t", "Cost to serve\nmedio R$/t",
         "Piso margem %", "1.Preco caixa", "2.Preco contrib.", "3.Preco equilibrio",
         "4.Preco min. autoriz.", "5.Preco-alvo", "6.Preco estrategico", "7.Preco praticado"]
    W = {"A": 11}
    for col in "BCDEFGHIJKL":
        W[col] = 12
    ws = custom_head("39_PRICING", "Commercial", "Motor de precificacao (7 niveis)",
                     "Cada nivel tem definicao matematica documentada. Caixa=custo incremental; "
                     "Contribuicao=custo variavel; Equilibrio=custo total+servir; Min autorizado=equilibrio x(1+piso).",
                     H, W, tab=COMM, span=12)
    matvar = "'27_ACTUAL_COST'!$D${a}:$D${b}".format(a=R0, b=R0 + 2)  # material act R$/t (proxy variavel)
    for i, p in enumerate(PRD):
        r = R0 + i
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, "=IFERROR(INDEX({m},MATCH(A{r},{a},0)),0)".format(m=matvar, a=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 3, "=IFERROR(INDEX({h},MATCH(A{r},{a},0)),0)".format(h=ACT_H, a=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 4, 5, style="input", fmt=MONEY, align="r")
        put(ws, r, 5, 0.20, style="input", fmt=PCT2, align="r")
        put(ws, r, 6, "=B{r}".format(r=r), style="form", fmt=MONEY, align="r")           # caixa=variavel incremental
        put(ws, r, 7, "=B{r}".format(r=r), style="form", fmt=MONEY, align="r")           # contribuicao=variavel
        put(ws, r, 8, "=C{r}+D{r}".format(r=r), style="form", fmt=MONEY, align="r")      # equilibrio=total+servir
        put(ws, r, 9, "=H{r}*(1+E{r})".format(r=r), style="form", fmt=MONEY, align="r")  # min autorizado
        put(ws, r, 10, None, style="input", fmt=MONEY, align="r")  # preco-alvo (target costing)
        put(ws, r, 11, None, style="input", fmt=MONEY, align="r")  # estrategico
        put(ws, r, 12, None, style="input", fmt=MONEY, align="r")  # praticado (link 33)
    return ws


# ===========================================================================
# 32_SAVINGS  (funnel + estagios + regra de realizado)
# ===========================================================================
def build_savings():
    cols = [
        C("SAVING_ID", 10, "key"), C("Causa", 18), C("Processo", 12), C("PRODUCT_ID", 11),
        C("Baseline R$/t", 12, fmt=MONEY), C("Target R$/t", 12, fmt=MONEY),
        C("Gap R$/t", 10, "form", fmt=MONEY), C("Volume t/ano", 12, fmt=INT),
        C("Saving bruto R$/ano", 14, "form", fmt=INT), C("Investimento R$", 12, fmt=MONEY),
        C("Custo implem. R$", 12, fmt=MONEY), C("Saving liq. R$/ano", 14, "form", fmt=INT),
        C("Estagio (1-8)", 12), C("Evidencia oper.?", 12), C("Confirm. financ.?", 12),
        C("Realizado?", 10, "form"), C("Priority score", 12, "form", fmt=NUM),
        C("Responsavel", 14), C("Prazo", 11, fmt="yyyy-mm-dd"),
    ]
    demo = [
        ["SAV-001", "Reduzir consumo cimento", "Formulacao", "P001", 160, 154, "=E7-F7", 60000,
         "=G7*H7", 5000, 2000, "=I7-J7-K7", "3.Validado tec.", "Nao", "Nao", '=IF(AND(N7="Sim",O7="Sim"),"SIM","NAO")', "=IFERROR(I7/1000*0.6*3/(J7/1000+1),0)", "Eng. Tecnico", "2026-06-30"],
        ["SAV-002", "Reduzir ociosidade MIST2", "Producao", "P003", 58, 52, "=E8-F8", 8000,
         "=G8*H8", 0, 500, "=I8-J8-K8", "2.Em analise", "Nao", "Nao", '=IF(AND(N8="Sim",O8="Sim"),"SIM","NAO")', "=IFERROR(I8/1000*0.5*4/(J8/1000+1),0)", "Producao", "2026-05-31"],
    ]
    table_sheet("32_SAVINGS", "Cost Reduction", "Motor de reducao de custos (savings funnel)",
                "8 estagios: Identificado->Sustentado. So e REALIZADO com evidencia operacional E "
                "confirmacao financeira E reconciliacao (sem transferir custo, sem perder qualidade). "
                "Priority score = impacto x prob x velocidade / investimento.",
                cols, demo=demo, tab="C00000", span=19)


# ===========================================================================
# 42_RECON_COST  &  43_RECON_ACCOUNTING
# ===========================================================================
def build_recon_cost():
    H = ["PRODUCT_ID", "Std total R$\n(26!G)", "PPV R$", "Usage R$", "Conv var R$",
         "Total var R$", "Std+Var R$", "Actual R$\n(27!G)", "Residual R$"]
    W = {"A": 11, "B": 13, "C": 11, "D": 11, "E": 11, "F": 11, "G": 13, "H": 13, "I": 11}
    ws = custom_head("42_RECON_COST", "Reconciliation", "Reconciliacao industrial (Std+Var=Actual)",
                     "Para cada produto: Custo padrao + variacoes = custo real. Residual deve -> 0. "
                     "Se residual !=0, ha variacao nao identificada (nunca esconder em 'Outros').",
                     H, W, tab="203864", span=9)
    vsa = "'28_VARIANCES'!$A${a}:$A${b}".format(a=VAR_SUM_START, b=VAR_SUM_START + 2)
    def vsum(col):
        return "IFERROR(INDEX('28_VARIANCES'!${c}${a}:${c}${b},MATCH(A{{r}},{sa},0)),0)".format(c=col, a=VAR_SUM_START, b=VAR_SUM_START + 2, sa=vsa)
    for i, p in enumerate(PRD):
        r = R0 + i
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, "=IFERROR(INDEX('26_STANDARD_COST'!$G${a}:$G${b},MATCH(A{r},{sa},0)),0)".format(a=R0, b=R0 + 2, sa=STD_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 3, "=" + vsum("D").format(r=r), style="form", fmt=MONEY, align="r")   # PPV
        put(ws, r, 4, "=" + vsum("E").format(r=r), style="form", fmt=MONEY, align="r")   # Usage
        put(ws, r, 5, "=" + vsum("K").format(r=r), style="form", fmt=MONEY, align="r")   # Conv var
        put(ws, r, 6, "=C{r}+D{r}+E{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, "=B{r}+F{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 8, "=IFERROR(INDEX('27_ACTUAL_COST'!$G${a}:$G${b},MATCH(A{r},{sa},0)),0)".format(a=R0, b=R0 + 2, sa=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 9, "=ROUND(G{r}-H{r},2)".format(r=r), style="control", fmt=MONEY4, align="r", bold=True)
    put(ws, R0 + 4, 1, "Residual total:", style="note", border=False)
    put(ws, R0 + 4, 9, "=ROUND(SUM(I{a}:I{b}),2)".format(a=R0, b=R0 + 2), style="control", fmt=MONEY4, align="r", bold=True)
    return ws


def build_recon_accounting():
    cols = [
        C("Conta contabil", 22, "input"), C("Centro custo", 12), C("Cost model R$", 14, "input", fmt=MONEY),
        C("Razao/CPV R$", 14, "input", fmt=MONEY), C("Diferenca R$", 12, "form", fmt=MONEY),
        C("Natureza da diferenca", 26, "input"),
    ]
    demo = [
        ["Materias-primas consumidas", "CC-PROD", None, None, "=C7-D7", "Competencia/estoque"],
        ["Mao de obra direta", "CC-PROD", None, None, "=C8-D8", "Rateio/provisao"],
        ["Energia", "CC-PROD", None, None, "=C9-D9", "Competencia"],
        ["Depreciacao", "CC-PROD", None, None, "=C10-D10", "Metodo"],
        ["Manutencao", "CC-MANUT", None, None, "=C11-D11", "Provisao"],
        ["UNALLOCATED (sem driver)", "-", None, None, "=C12-D12", "Custo nao alocado por falta de driver"],
    ]
    table_sheet("43_RECON_ACCOUNTING", "Reconciliation", "Ponte custo industrial <-> contabilidade (CPV)",
                "Bridge entre COST MODEL e RAZAO/CPV: contas, centros, estoque, provisoes, depreciacao, "
                "competencia, rateios. Sem driver causal defensavel -> UNALLOCATED (nao ratear).",
                cols, demo=demo, tab="203864", span=6)


def kpi_tile(ws, r, c, label, formula, fmt=MONEY, span=2, fill=FILL_LINKED, alert=False):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    lc = ws.cell(r, c, label); lc.font = F_NOTE; lc.fill = fill; lc.alignment = AL_L
    lc.border = BORDER_ALL
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + span - 1)
    vc = ws.cell(r + 1, c, formula)
    vc.font = Font(name="Calibri", size=14, bold=True, color=("C00000" if alert else NAVY))
    vc.fill = fill; vc.alignment = AL_R; vc.number_format = fmt; vc.border = BORDER_ALL
    ws.row_dimensions[r + 1].height = 24
    for cc in range(c, c + span):
        ws.cell(r, cc).border = BORDER_ALL
        ws.cell(r + 1, cc).border = BORDER_ALL


def build_dashboards():
    # aggregate ranges
    of_tot = "SUM('17_PRODUCTION_OF'!$L${a}:$L${b})".format(a=R0, b=OF_LAST)
    act_tot = "SUM('27_ACTUAL_COST'!$G${a}:$G${b})".format(a=R0, b=R0 + 2)
    std_tot = "SUM('26_STANDARD_COST'!$G${a}:$G${b})".format(a=R0, b=R0 + 2)
    vs = (VAR_SUM_START, VAR_SUM_START + 2)
    ppv = "SUM('28_VARIANCES'!$D${a}:$D${b})".format(a=vs[0], b=vs[1])
    usage = "SUM('28_VARIANCES'!$E${a}:$E${b})".format(a=vs[0], b=vs[1])
    yield_ = "SUM('28_VARIANCES'!$I${a}:$I${b})".format(a=vs[0], b=vs[1])
    mix = "SUM('28_VARIANCES'!$J${a}:$J${b})".format(a=vs[0], b=vs[1])
    conv = "SUM('28_VARIANCES'!$K${a}:$K${b})".format(a=vs[0], b=vs[1])
    unused = "SUM('29_CAPACITY'!$F${a}:$F${b})".format(a=R0, b=R0 + 1)
    copq = "SUM('24_QUALITY'!$E${a}:$E${b})".format(a=R0, b=TB)
    sav_col = ("'32_SAVINGS'!$P${a}:$P${b}".format(a=R0, b=TB), "'32_SAVINGS'!$L${a}:$L${b}".format(a=R0, b=TB))
    sav_real = 'SUMIF({p},"SIM",{v})'.format(p=sav_col[0], v=sav_col[1])
    sav_pipe = 'SUMIF({p},"NAO",{v})'.format(p=sav_col[0], v=sav_col[1])

    # ---- 44_EXEC_DASHBOARD ----
    ws = sheet("44_EXEC_DASHBOARD", "Dashboard", "Painel executivo do proprietario", tab=NAVY)
    widths(ws, {"A": 3, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20})
    title(ws, "PAINEL EXECUTIVO — SKAL COST INTELLIGENCE",
          "Poucos numeros, alto impacto. Todos os valores sao calculados das abas de origem. "
          "R$/ton boa produzida como unidade base.", span=7)
    kpi_tile(ws, 4, 2, "CUSTO REAL R$/t", "=IFERROR({a}/{t},0)".format(a=act_tot, t=of_tot))
    kpi_tile(ws, 4, 4, "CUSTO PADRAO R$/t", "=IFERROR({s}/{t},0)".format(s=std_tot, t=of_tot))
    kpi_tile(ws, 4, 6, "COST GAP TOTAL R$", "={a}-{s}".format(a=act_tot, s=std_tot), alert=True)
    kpi_tile(ws, 7, 2, "COST GAP R$/t", "=IFERROR(({a}-{s})/{t},0)".format(a=act_tot, s=std_tot, t=of_tot), alert=True)
    kpi_tile(ws, 7, 4, "SAVINGS REALIZADOS R$", "=" + sav_real, fill=FILL_CONTROL)
    kpi_tile(ws, 7, 6, "SAVINGS PIPELINE R$", "=" + sav_pipe, fill=FILL_CONTROL)
    kpi_tile(ws, 10, 2, "UNUSED CAPACITY R$", "=" + unused, alert=True)
    kpi_tile(ws, 10, 4, "COPQ R$", "=" + copq, alert=True)
    kpi_tile(ws, 10, 6, "GOOD TONS", "=" + of_tot, fmt=NUM)
    # variance bridge
    r = 14
    c = ws.cell(r, 2, "DECOMPOSICAO DO DESVIO (Actual - Standard)  —  R$"); c.font = F_SECT; r += 1
    header_row(ws, r, ["Componente", "R$", "% do gap"], start=2); r += 1
    comps = [("PPV (preco de compra)", ppv), ("Usage (consumo)", usage),
             ("  - Mix", mix), ("  - Yield", yield_), ("Conversao", conv)]
    gapf = "({a}-{s})".format(a=act_tot, s=std_tot)
    for lbl, f in comps:
        put(ws, r, 2, lbl, style="form", align="l")
        put(ws, r, 3, "=" + f, style="form", fmt=MONEY, align="r")
        put(ws, r, 4, "=IFERROR(C{r}/{g},0)".format(r=r, g=gapf), style="form", fmt=PCT, align="r")
        r += 1
    put(ws, r, 2, "TOTAL (PPV+Usage+Conversao)", style="control", align="l", bold=True)
    put(ws, r, 3, "={p}+{u}+{c}".format(p=ppv, u=usage, c=conv), style="control", fmt=MONEY, align="r", bold=True)
    put(ws, r, 4, "=IFERROR(C{r}/{g},0)".format(r=r, g=gapf), style="control", fmt=PCT, align="r")
    r += 2
    put(ws, r, 2, "Nota: PPV+Usage+Conversao deve igualar o Cost Gap Total (Actual-Standard). "
        "Ver 42_RECON_COST (residual -> 0).", style="note", border=False)
    ws.freeze_panes = "A4"

    # ---- generic linked-metric dashboards ----
    def metric_dash(name, purpose, sub, rows):
        w = sheet(name, "Dashboard", purpose, tab=NAVY)
        widths(w, {"A": 3, "B": 40, "C": 20, "D": 40})
        title(w, name.split("_", 1)[1].replace("_", " "), sub, span=4)
        header_row(w, 4, ["Metrica", "Valor", "Fonte"], start=2)
        rr = 5
        for lbl, f, fmt, src in rows:
            put(w, rr, 2, lbl, style="form", align="l")
            if f is None:
                put(w, rr, 3, "DADO NECESSARIO", style="alert", align="c")
            else:
                put(w, rr, 3, "=" + f, style="link", fmt=fmt, align="r")
            put(w, rr, 4, src, style="note", align="l")
            rr += 1
        w.freeze_panes = "A4"

    metric_dash("45_PRODUCTION_DASHBOARD", "Painel de producao",
                "Produtividade, yield, perdas, energia, COPQ, custo controlavel.",
                [("Good tons (mes)", of_tot, NUM, "17_PRODUCTION_OF"),
                 ("t/h medio (RES-MIST1)", "IFERROR(SUMIF('17_PRODUCTION_OF'!$E${a}:$E${b},\"RES-MIST1\",'17_PRODUCTION_OF'!$L${a}:$L${b})/SUMIF('17_PRODUCTION_OF'!$E${a}:$E${b},\"RES-MIST1\",'17_PRODUCTION_OF'!$Q${a}:$Q${b}),0)".format(a=R0, b=OF_LAST), NUM, "17"),
                 ("Perda total (t)", "SUM('17_PRODUCTION_OF'!$O${a}:$O${b})".format(a=R0, b=OF_LAST), NUM, "17"),
                 ("Nao explicado (t)", "SUM('19_LOSSES'!$I${a}:$I${b})".format(a=R0, b=TB), NUM, "19_LOSSES"),
                 ("Usage variance R$", usage, MONEY, "28"),
                 ("Yield variance R$", yield_, MONEY, "28"),
                 ("Energia R$/t", "IFERROR(SUM('21_ENERGY'!$H${a}:$H${b})/{t},0)".format(a=R0, b=TB, t=of_tot), MONEY, "21_ENERGY"),
                 ("COPQ R$", copq, MONEY, "24_QUALITY"),
                 ("Unused capacity R$", unused, MONEY, "29_CAPACITY")])

    metric_dash("46_PROCUREMENT_DASHBOARD", "Painel de compras",
                "PPV, landed cost, TCO, should cost, supplier gap, credito tributario.",
                [("PPV total R$", ppv, MONEY, "28"),
                 ("Landed net Cimento R$/t", "IFERROR(INDEX('09_LANDED_COST'!$P${a}:$P${b},MATCH(\"M001\",'09_LANDED_COST'!$A${a}:$A${b},0)),0)".format(a=R0, b=LANDED_LAST), MONEY, "09"),
                 ("Credito tributario potencial R$", "SUM('11_TAX_CREDITS'!$D${a}:$D${b})+SUM('11_TAX_CREDITS'!$E${a}:$E${b})".format(a=R0, b=TB), MONEY, "11_TAX_CREDITS"),
                 ("Should cost gap R$", None, MONEY, "16 (DATA_GAPS: preco mercado/should cost)"),
                 ("Supplier TCO ranking", None, MONEY, "15/TCO (DATA_GAPS)")])

    total_margin = "SUM('40_PROFITABILITY'!$R${a}:$R${b})".format(a=R0, b=R0 + 2)
    metric_dash("47_COMMERCIAL_DASHBOARD", "Painel comercial",
                "Preco, pocket price, cost to serve, margem, cliente x SKU. "
                "O comercial NAO altera formulas de custo industrial.",
                [("Pocket margin total R$", total_margin, MONEY, "40_PROFITABILITY"),
                 ("Cost to serve total R$", "SUM('38_COST_TO_SERVE'!$K${a}:$K${b})".format(a=R0, b=TB), MONEY, "38"),
                 ("Vendas total t", "SUM('33_SALES'!$E${a}:$E${b})".format(a=R0, b=TB), NUM, "33_SALES"),
                 ("Destruidores de valor (qtd)", "COUNTIF('40_PROFITABILITY'!$S${a}:$S${b},\"DESTRUIDOR DE VALOR\")".format(a=R0, b=R0 + 2), INT, "40"),
                 ("Inadimplencia esperada R$", "SUM('36_RECEIVABLES'!$G${a}:$G${b})".format(a=R0, b=TB), MONEY, "36")])

    metric_dash("48_COST_REDUCTION_DASHBOARD", "Painel de reducao de custos",
                "Potencial identificado -> validado -> implementado -> realizado -> sustentado.",
                [("Saving bruto (pipeline) R$/ano", "SUM('32_SAVINGS'!$I${a}:$I${b})".format(a=R0, b=TB), INT, "32_SAVINGS"),
                 ("Saving liquido (pipeline) R$/ano", "SUM('32_SAVINGS'!$L${a}:$L${b})".format(a=R0, b=TB), INT, "32"),
                 ("Saving REALIZADO R$", sav_real, MONEY, "32 (regra: evidencia+financeiro)"),
                 ("Cost gap controlavel R$/ano", "SUMIF('31_COST_GAP'!$M${a}:$M${b},\"<>\")".format(a=R0, b=R0 + 2), INT, "31_COST_GAP"),
                 ("Top priority score", "IFERROR(MAX('32_SAVINGS'!$Q${a}:$Q${b}),0)".format(a=R0, b=TB), NUM, "32")])


# ===========================================================================
# DATA_DICTIONARY
# ===========================================================================
def build_data_dictionary():
    cols = [C("Campo", 20, "form"), C("Tabela/Aba", 20, "form"), C("Tipo", 10, "form"),
            C("Unidade", 10, "form"), C("Obrigatorio", 10, "form"), C("Fonte", 14, "form"),
            C("Definicao", 44, "form")]
    demo = [
        ["PRODUCT_ID", "03_PRODUCTS", "texto", "-", "Sim", "Cadastro", "Chave unica do produto (SKU)"],
        ["MATERIAL_ID", "04_MATERIALS", "texto", "-", "Sim", "Cadastro", "Chave unica da materia-prima"],
        ["Extras?", "03_PRODUCTS", "bool", "-", "Sim", "Producao", "TRUE -> 20kg/saco,75/palete; FALSE -> 15/100"],
        ["kg/t produto", "06_BOM", "numero", "kg/t", "Sim", "Tecnico", "Quantidade padrao de MP por tonelada de produto"],
        ["LANDED NET R$/t", "09_LANDED_COST", "numero", "R$/t", "Sim", "Compras/Fiscal", "Custo liquido posto fabrica por tonelada"],
        ["Good tons", "17_PRODUCTION_OF", "numero", "t", "Sim", "Producao", "Toneladas boas liberadas (base de todos os R$/t)"],
        ["Act kg/t", "18_ACTUAL_CONSUMPTION", "numero", "kg/t", "Sim", "Producao", "Consumo real por tonelada boa"],
        ["PPV", "28_VARIANCES", "numero", "R$", "Calc", "Sistema", "(Preco real - padrao) x qtd real"],
        ["Usage Variance", "28_VARIANCES", "numero", "R$", "Calc", "Sistema", "(Qtd real - padrao permitida) x preco padrao"],
        ["Pocket Price", "40_PROFITABILITY", "numero", "R$/t", "Calc", "Sistema", "Preco liquido de bolso (waterfall)"],
        ["Saving liquido", "32_SAVINGS", "numero", "R$/ano", "Calc", "Sistema", "Saving bruto - investimento - custo implementacao"],
        ["ICMS_EFETIVO", "01_CONTROL_PANEL", "numero", "%", "Sim", "Fiscal", "ICMS nominal x % efetivamente devido (benef. 80%)"],
    ]
    table_sheet("DATA_DICTIONARY", "Doc", "Dicionario de dados (campos e definicoes)",
                "Nenhum campo ambiguo. Chaves unicas nunca dependem apenas da descricao textual.",
                cols, demo=demo, tab="1F3864", span=7)


# ===========================================================================
# DATA_GAPS
# ===========================================================================
def build_data_gaps():
    cols = [C("Informacao faltante", 34, "form"), C("Impacto", 30, "form"),
            C("Responsavel", 14, "form"), C("Prioridade", 10, "form"),
            C("Finalidade", 26, "form"), C("Formato necessario", 22, "form")]
    demo = [
        ["Planilha_Precificacao.xlsx (arquivo legado)", "Sem ela nao ha reconciliacao com o modelo atual (LEGACY_RECON)", "Controladoria", "ALTA", "Mapear precos/formulas atuais", ".xlsx"],
        ["CustoFormula2026.1.xlsb", "Estrutura detalhada de custos atual", "Controladoria", "ALTA", "Reconciliacao e validacao", ".xlsb"],
        ["Landed cost real das MP de PRODUCAO (areia/aditivos)", "Cimento/calcario ja reais (R3); faltam demais MP", "Compras/Fiscal", "ALTA", "R3_LANDED_REAL / 09", "Itens NF por MP"],
        ["Frete inbound do calcario e demais MP", "Landed do calcario incompleto (so cimento tem frete)", "Compras/Log.", "MEDIA", "R3_LANDED_REAL", "CT-e por carga"],
        ["Vinculo NF-item x OF de producao", "Ligar compra real ao consumo/BOM", "Producao/Compras", "ALTA", "R2 -> 18", "Chave OF"],
        ["kWh por equipamento", "Energia/t estimada", "Producao", "ALTA", "21_ENERGY", "kWh por OF/recurso"],
        ["Preco historico por fornecedor", "PPV / should cost incompletos", "Compras", "ALTA", "10/PPV", "Serie temporal"],
        ["Perda e reprocesso por OF", "Yield e mass balance incompletos", "Producao", "ALTA", "17/19/20", "Por OF"],
        ["Tempos reais de inicio/fim por lote/traco", "Produtividade real (t/h) por OF", "Producao", "ALTA", "17_PRODUCTION_OF", "hh:mm por OF"],
        ["Custos fixos por recurso (dep/manut/MO)", "R$/hora e unused capacity", "Controladoria", "MEDIA", "12/15/29", "R$/mes por recurso"],
        ["Entitlement / best demonstrated / theoretical", "Cost gap completo", "Eng./Controladoria", "MEDIA", "31_COST_GAP", "R$/t por SKU"],
        ["Preco de mercado / should cost", "Procurement gap", "Compras", "MEDIA", "16/46", "R$/un"],
        ["Base de vendas cliente x SKU completa", "Rentabilidade e pocket margin", "Comercial", "ALTA", "33/40", "Por venda"],
        ["Inadimplencia historica (PD/LGD)", "Expected credit loss", "Financeiro", "MEDIA", "36", "% por cliente/faixa"],
        ["Razao contabil / CPV por conta", "Reconciliacao contabil", "Contabilidade", "ALTA", "43", "Balancete"],
        ["Parametros SUDENE documentados", "IRPJ liquido correto", "Fiscal", "MEDIA", "09/tax", "Laudo/ADE"],
        ["Aliquotas reforma (CBS/IBS) por ano", "Cenarios 2026-2033", "Fiscal", "MEDIA", "10_TAX_RULES", "% por vigencia"],
    ]
    table_sheet("DATA_GAPS", "Doc", "Relatorio de lacunas de dados",
                "Cada linha e um campo de entrada ja criado no sistema, aguardando dado real. "
                "Prioridade por R$ de impacto anual (ver 01_CONTROL_PANEL materialidade).",
                cols, demo=demo, tab="C00000", span=6)


# ===========================================================================
# LEGACY_RECON
# ===========================================================================
def build_legacy_recon():
    H = ["PRODUCT_ID", "Preco atual\n(legado)", "Custo atual\n(legado)", "Novo custo\nindustrial R$/t",
         "Novo custo\neconomico R$/t", "Dif. abs.\nR$/t", "Dif. %", "Causa da diferenca"]
    W = {"A": 11, "B": 12, "C": 12, "D": 13, "E": 13, "F": 11, "G": 9, "H": 30}
    ws = custom_head("LEGACY_RECON", "Doc", "Reconciliacao modelo legado vs novo",
                     "Comparacao produto a produto. Legado (amarelo) vem da Planilha_Precificacao.xlsx "
                     "(ver DATA_GAPS). Novo custo industrial e calculado (27_ACTUAL_COST).", H, W, tab="1F3864", span=8)
    for i, p in enumerate(PRD):
        r = R0 + i
        put(ws, r, 1, p, style="key", align="c")
        put(ws, r, 2, None, style="input", fmt=MONEY, align="r")  # preco legado
        put(ws, r, 3, None, style="input", fmt=MONEY, align="r")  # custo legado
        put(ws, r, 4, "=IFERROR(INDEX({h},MATCH(A{r},{a},0)),0)".format(h=ACT_H, a=ACT_A, r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 5, "=D{r}".format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 6, '=IF(C{r}="","",D{r}-C{r})'.format(r=r), style="form", fmt=MONEY, align="r")
        put(ws, r, 7, '=IF(C{r}="","",IFERROR((D{r}-C{r})/C{r},0))'.format(r=r), style="form", fmt=PCT2, align="r")
        put(ws, r, 8, None, style="input", align="l")
    put(ws, R0 + 4, 1, "Preencha as colunas do legado ao receber Planilha_Precificacao.xlsx. "
        "Causas tipicas: divergencia de metodologia, premissa sem suporte, dado nao rastreavel.",
        style="note", border=False)
    return ws


# ===========================================================================
# CONTROLS_ALERTS  (sistema de alertas automaticos)
# ===========================================================================
def build_alerts():
    ws = sheet("CONTROLS_ALERTS", "Doc", "Alertas e testes automaticos do modelo", tab="C00000")
    widths(ws, {"A": 3, "B": 44, "C": 18, "D": 40})
    title(ws, "CONTROLES E ALERTAS AUTOMATICOS",
          "Verde = OK, Vermelho = investigar. Estes testes implementam a aceitacao do modelo "
          "(BOM fecha, Std+Var=Actual, sem custo negativo, etc.).", span=4)
    header_row(ws, 4, ["Verificacao", "Status", "Regra"], start=2)
    r = 5
    checks = [
        ("BOM fecha (soma kg/t = 1000) — P001",
         '=IF(ABS(SUMIF(\'06_BOM\'!$A${a}:$A${b},"P001",\'06_BOM\'!$E${a}:$E${b})-1000)<=1,"OK","ALERTA")'.format(a=R0, b=BOM_LASTR),
         "Fechamento fisico da formulacao"),
        ("BOM fecha — P002",
         '=IF(ABS(SUMIF(\'06_BOM\'!$A${a}:$A${b},"P002",\'06_BOM\'!$E${a}:$E${b})-1000)<=1,"OK","ALERTA")'.format(a=R0, b=BOM_LASTR),
         "Fechamento fisico"),
        ("BOM fecha — P003",
         '=IF(ABS(SUMIF(\'06_BOM\'!$A${a}:$A${b},"P003",\'06_BOM\'!$E${a}:$E${b})-1000)<=1,"OK","ALERTA")'.format(a=R0, b=BOM_LASTR),
         "Fechamento fisico"),
        ("Reconciliacao Std+Var=Actual (residual total)",
         '=IF(ABS(SUM(\'28_VARIANCES\'!$N${a}:$N${b}))<0.01,"OK","ALERTA")'.format(a=R0, b=VAR_LAST),
         "Residual por linha -> 0"),
        ("Reconciliacao por produto (42_RECON_COST)",
         '=IF(ABS(SUM(\'42_RECON_COST\'!$I${a}:$I${b}))<0.01,"OK","ALERTA")'.format(a=R0, b=R0 + 2),
         "Std total + var = Actual"),
        ("Sem custo real negativo",
         '=IF(MIN(\'27_ACTUAL_COST\'!$H${a}:$H${b})>=0,"OK","ALERTA")'.format(a=R0, b=R0 + 2),
         "Custo/t nao pode ser negativo"),
        ("Material sem landed cost (custo zero no BOM)",
         '=IF(SUMPRODUCT((\'06_BOM\'!$E${a}:$E${b}>0)*(\'06_BOM\'!$K${a}:$K${b}=0))=0,"OK","ALERTA")'.format(a=R0, b=BOM_LASTR),
         "Toda MP usada deve ter custo rastreavel"),
        ("Perda nao explicada (mass balance)",
         '=IF(SUM(\'19_LOSSES\'!$I${a}:$I${b})<=0.5,"OK","ALERTA")'.format(a=R0, b=TB),
         "Unaccounted loss nao absorvida"),
        ("Produtos destruidores de valor",
         '=IF(COUNTIF(\'40_PROFITABILITY\'!$S${a}:$S${b},"DESTRUIDOR DE VALOR")=0,"OK","ALERTA")'.format(a=R0, b=R0 + 2),
         "Pocket margin < 0"),
        ("ICMS efetivo consistente (nominal x devido)",
         '=IF(ABS(ICMS_EFETIVO-ICMS_NOMINAL*(1-ICMS_REDUCAO))<0.0001,"OK","ALERTA")',
         "Sem hardcode de 4,5%"),
    ]
    for lbl, f, rule in checks:
        put(ws, r, 2, lbl, style="form", align="l")
        cell = put(ws, r, 3, f, style="control", align="c", bold=True)
        put(ws, r, 4, rule, style="note", align="l")
        r += 1
    put(ws, r + 1, 2, "TESTES DE ACEITACAO (98): rastreabilidade de MP, BOM fecha, OF reconcilia, "
        "Std+Var=Actual, ponte contabil, sem alocacao sem driver, fonte+responsavel, "
        "preco decomponivel ate custo, saving reconciliavel, leitura em segundos.", style="note", border=False)
    ws.freeze_panes = "A4"
    return ws


# ===========================================================================
# REAL DATA LAYER — dados reais das notas de compra (ingest_compras.py)
# ===========================================================================
def build_real_procurement():
    try:
        import ingest_compras
        R = ingest_compras.load_all("data")
    except Exception as e:
        print("  [real] ingest indisponivel:", e)
        R = None
    if not R or not R.get("files"):
        return
    files = R["files"]
    fonte_txt = "Fonte: " + ", ".join("%s (%s)" % (m, f) for f, m in files)

    # ranking de gasto autoritativo = agregacao do ledger de documentos (valor por doc),
    # que e mais completo que a aba Fornecedores de cada mes. Chave: CNPJ (fallback nome).
    import re as _re
    def _norm(nm):
        if not nm:
            return None
        s = _re.sub(r"[^A-Z0-9 ]", " ", str(nm).upper())
        for suf in (" LTDA", " S A", " SA", " ME", " EIRELI", " EPP", " EI",
                    " INDUSTRIA", " COMERCIO", " PRESTADORA", " DE SERVICOS"):
            s = s.replace(suf, " ")
        s = _re.sub(r"\s+", " ", s).strip()
        return s[:24] or None

    meta = {}  # cnpj/nome -> {name, uf, regime}
    name2cnpj = {}
    for s in R["suppliers"]:
        meta[s["cnpj"] or s["name"]] = s
        if s.get("cnpj") and _norm(s["name"]):
            name2cnpj[_norm(s["name"])] = s["cnpj"]
    for d in R["docs"]:            # aprende nome->cnpj tambem do ledger
        if d.get("cnpj") and _norm(d.get("emit")):
            name2cnpj.setdefault(_norm(d.get("emit")), d["cnpj"])
    spend = {}
    for d in R["docs"]:
        key = d.get("cnpj") or name2cnpj.get(_norm(d.get("emit"))) or d.get("emit")
        if not key:
            continue
        e = spend.setdefault(key, {"cnpj": d.get("cnpj"), "name": None,
                                   "total": 0.0, "months": set(), "docs": 0})
        e["total"] += d.get("valor") or 0.0
        e["months"].add(d["month"])
        e["docs"] += 1
        if not e["name"] and d.get("emit"):
            e["name"] = d.get("emit")
    for key, e in spend.items():
        m = meta.get(key) or meta.get(e.get("name"))
        e["uf"] = (m or {}).get("uf")
        e["regime"] = (m or {}).get("regime")
        if not e["name"]:
            e["name"] = (m or {}).get("name") or (str(key))
    real_suppliers = sorted(spend.values(), key=lambda x: -x["total"]) or R["suppliers"]

    # ---- R0: sumario real ----
    ws = sheet("R0_DADOS_REAIS", "Real", "Sumario dos dados reais de compras", tab="C55A11")
    widths(ws, {"A": 3, "B": 34, "C": 20, "D": 20, "E": 40})
    title(ws, "DADOS REAIS — COMPRAS (NOTAS DE ENTRADA)",
          fonte_txt + ". Itens sem quantidade/valor legiveis no OCR foram descartados "
          "(nunca inferidos). Estes numeros SAO reais, ao contrario do dataset DEMO.", span=5)
    r = 4
    tot_sup = sum(s["total"] for s in real_suppliers)
    stats = [("Arquivos ingeridos", len(files)),
             ("Fornecedores (dedup CNPJ)", len(real_suppliers)),
             ("Gasto total mapeado (R$)", round(tot_sup, 2)),
             ("Itens de produto legiveis", len(R["product_items"])),
             ("Documentos no ledger", len(R["docs"])),
             ("CT-e de frete", len(R["freights"])),
             ("Registros de energia", len(R["energy"]))]
    for lbl, v in stats:
        put(ws, r, 2, lbl, style="form", align="l")
        put(ws, r, 3, v, style="link", fmt=(MONEY if "R$" in lbl else INT), align="r")
        r += 1
    r += 1
    put(ws, r, 2, "ACHADO-CHAVE", style="form", bold=True, align="l")
    put(ws, r, 5, "O frete inbound do cimento (Apodi/CE -> Teresina/PI) e um driver de custo "
        "de primeira ordem: ver R3_LANDED_REAL.", style="alert", align="l")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=5)
    ws.freeze_panes = "A4"

    # ---- R1: fornecedores reais ----
    wf = sheet("R1_FORNECEDORES_REAL", "Real", "Fornecedores reais (ranking de gasto)", tab="C55A11")
    widths(wf, {"A": 18, "B": 42, "C": 6, "D": 22, "E": 14, "F": 12})
    title(wf, "FORNECEDORES REAIS", fonte_txt +
          " Gasto = soma dos documentos (ledger R4), agregado por CNPJ.", span=6)
    header_row(wf, 3, ["CNPJ", "Razao social", "UF", "Regime", "Gasto total R$", "Meses"])
    rr = 4
    for s in real_suppliers:
        put(wf, rr, 1, s["cnpj"], style="key", align="c")
        put(wf, rr, 2, s["name"], style="form", align="l")
        put(wf, rr, 3, s.get("uf"), style="form", align="c")
        put(wf, rr, 4, s.get("regime"), style="form", align="l")
        put(wf, rr, 5, round(s["total"], 2), style="link", fmt=MONEY, align="r")
        put(wf, rr, 6, ", ".join(sorted(s["months"])), style="form", align="c")
        rr += 1
    put(wf, rr, 4, "TOTAL", style="control", bold=True, align="r")
    put(wf, rr, 5, "=SUM(E4:E%d)" % (rr - 1), style="control", fmt=MONEY, align="r", bold=True)
    wf.freeze_panes = "A4"

    # ---- R2: compras reais (itens de produto legiveis) ----
    wp = sheet("R2_COMPRAS_REAIS", "Real", "Compras reais por item (OCR legivel)", tab="C55A11")
    widths(wp, {"A": 8, "B": 10, "C": 40, "D": 12, "E": 6, "F": 10, "G": 12, "H": 12, "I": 12})
    title(wp, "COMPRAS REAIS POR ITEM", fonte_txt +
          " Somente itens com quantidade e valor unitario numericos.", span=9)
    header_row(wp, 3, ["Mes", "Doc", "Descricao", "NCM", "Un", "Qtd", "Valor unit. R$", "Valor total R$", "Emitente"])
    rr = 4
    for it in R["product_items"]:
        put(wp, rr, 1, it["month"], style="key", align="c")
        put(wp, rr, 2, it["doc"], style="form", align="c")
        put(wp, rr, 3, it["desc"], style="form", align="l")
        put(wp, rr, 4, it["ncm"], style="form", align="c")
        put(wp, rr, 5, it["un"], style="form", align="c")
        put(wp, rr, 6, it["qty"], style="link", fmt=NUM, align="r")
        put(wp, rr, 7, it["vunit"], style="link", fmt=MONEY, align="r")
        put(wp, rr, 8, it["vtot"], style="form", fmt=MONEY, align="r")
        put(wp, rr, 9, it["emit"], style="form", align="l")
        rr += 1
    wp.freeze_panes = "A4"

    # ---- R3: landed cost real (com frete inbound) ----
    # cimento agosto
    def agg(pred, key_month="2026-08"):
        tons = sum(i["qty"] for i in R["product_items"] if pred(i) and i["month"] == key_month)
        val = sum(i["vtot"] for i in R["product_items"] if pred(i) and i["month"] == key_month)
        return tons, val
    cim_t, cim_v = agg(lambda i: "CIMENTO" in (i["desc"] or "").upper())
    cal_t, cal_v = agg(lambda i: "CALCARIO" in (i["desc"] or "").upper() or "CALCARIO" in (i["desc"] or "").upper())
    frete_cim = sum(f["valor"] or 0 for f in R["freights"]
                    if f["month"] == "2026-08" and "APODI" in (f["remetente"] or "").upper())
    wl = sheet("R3_LANDED_REAL", "Real", "Custo real posto fabrica (landed) — com frete inbound", tab="C55A11")
    widths(wl, {"A": 3, "B": 30, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})
    title(wl, "LANDED COST REAL — AGOSTO 2026",
          "Preco de NF NAO e custo. O frete inbound entra no custo posto fabrica. "
          "Cimento CST 500 = ICMS por substituicao (sem credito na entrada).", span=7)
    header_row(wl, 4, ["", "Material", "Qtd (t)", "NF R$/t", "Frete R$/t", "LANDED R$/t", "% frete s/ NF"], start=1)
    rr = 5
    # cimento row
    put(wl, rr, 2, "Cimento CP-V ARI (Apodi)", style="form", align="l")
    put(wl, rr, 3, round(cim_t, 2), style="link", fmt=NUM, align="r")
    put(wl, rr, 4, round(cim_v / cim_t, 2) if cim_t else 0, style="link", fmt=MONEY, align="r")
    put(wl, rr, 5, round(frete_cim / cim_t, 2) if cim_t else 0, style="link", fmt=MONEY, align="r")
    put(wl, rr, 6, "=D%d+E%d" % (rr, rr), style="form", fmt=MONEY, align="r", bold=True)
    put(wl, rr, 7, "=IFERROR(E%d/D%d,0)" % (rr, rr), style="form", fmt=PCT, align="r")
    rr += 1
    put(wl, rr, 2, "Calcario dolomitico #170", style="form", align="l")
    put(wl, rr, 3, round(cal_t, 2), style="link", fmt=NUM, align="r")
    put(wl, rr, 4, round(cal_v / cal_t, 2) if cal_t else 0, style="link", fmt=MONEY, align="r")
    put(wl, rr, 5, None, style="input", fmt=MONEY, align="r")  # frete calcario a confirmar
    put(wl, rr, 6, "=D%d+IF(E%d=\"\",0,E%d)" % (rr, rr, rr), style="form", fmt=MONEY, align="r", bold=True)
    put(wl, rr, 7, '=IFERROR(IF(E%d="",0,E%d)/D%d,0)' % (rr, rr, rr), style="form", fmt=PCT, align="r")
    rr += 2
    put(wl, rr, 2, "Frete cimento total (Apodi, ago): R$ %.2f  em %d t  =>  R$ %.2f/t"
        % (frete_cim, round(cim_t, 2), (frete_cim / cim_t if cim_t else 0)), style="note", border=False)
    rr += 1
    put(wl, rr, 2, "Leitura: a NF do cimento e R$505/t, mas o LANDED e ~R$%.0f/t. O frete inbound "
        "adiciona ~%.0f%% ao custo. Calcario: confirmar frete (DATA_GAPS)."
        % ((cim_v / cim_t + frete_cim / cim_t) if cim_t else 0,
           (frete_cim / cim_v * 100) if cim_v else 0), style="note", border=False)
    wl.freeze_panes = "A5"

    # ---- R4: ledger de notas ----
    wd = sheet("R4_NOTAS_LEDGER", "Real", "Ledger de notas de entrada (documento a documento)", tab="C55A11")
    widths(wd, {"A": 8, "B": 8, "C": 10, "D": 10, "E": 12, "F": 40, "G": 18, "H": 14, "I": 10, "J": 12})
    title(wd, "NOTAS DE ENTRADA — LEDGER", fonte_txt +
          " Nivel documento (Tributos_por_Doc): valor e ICMS por nota.", span=10)
    header_row(wd, 3, ["Mes", "Doc", "Tipo", "Numero", "Unidade", "Emitente", "CNPJ",
                       "Valor doc R$", "Aliq ICMS", "ICMS R$"])
    rr = 4
    tot_doc = 0.0
    for d in R["docs"]:
        put(wd, rr, 1, d["month"], style="key", align="c")
        put(wd, rr, 2, d["doc"], style="form", align="c")
        put(wd, rr, 3, d["tipo"], style="form", align="c")
        put(wd, rr, 4, d["num"], style="form", align="c")
        put(wd, rr, 5, d["unidade"], style="form", align="c")
        put(wd, rr, 6, d["emit"], style="form", align="l")
        put(wd, rr, 7, d["cnpj"], style="form", align="c")
        put(wd, rr, 8, d["valor"], style="link", fmt=MONEY, align="r")
        put(wd, rr, 9, d["aliq_icms"], style="form", fmt=PCT2, align="r")
        put(wd, rr, 10, d["icms"], style="form", fmt=MONEY, align="r")
        tot_doc += d["valor"] or 0
        rr += 1
    put(wd, rr, 6, "TOTAL (soma bruta, inclui servicos/energia/frete)", style="control", bold=True, align="r")
    put(wd, rr, 8, "=SUM(H4:H%d)" % (rr - 1), style="control", fmt=MONEY, align="r", bold=True)
    wd.freeze_panes = wd.cell(4, 1)

    # ---- R5: fretes ----
    wfr = sheet("R5_FRETES_CTE", "Real", "Fretes inbound (CT-e)", tab="C55A11")
    widths(wfr, {"A": 8, "B": 8, "C": 30, "D": 14, "E": 14, "F": 30, "G": 14})
    title(wfr, "FRETES INBOUND (CT-e)", fonte_txt, span=7)
    header_row(wfr, 3, ["Mes", "CT-e", "Transportadora", "Origem", "Destino", "Remetente", "Valor R$"])
    rr = 4
    for f in R["freights"]:
        put(wfr, rr, 1, f["month"], style="key", align="c")
        put(wfr, rr, 2, f["cte"], style="form", align="c")
        put(wfr, rr, 3, f["transp"], style="form", align="l")
        put(wfr, rr, 4, f["origem"], style="form", align="c")
        put(wfr, rr, 5, f["destino"], style="form", align="c")
        put(wfr, rr, 6, f["remetente"], style="form", align="l")
        put(wfr, rr, 7, f["valor"], style="link", fmt=MONEY, align="r")
        rr += 1
    put(wfr, rr, 6, "TOTAL", style="control", bold=True, align="r")
    put(wfr, rr, 7, "=SUM(G4:G%d)" % (rr - 1), style="control", fmt=MONEY, align="r", bold=True)
    wfr.freeze_panes = wfr.cell(4, 1)
    print("  [real] %d fornecedores, %d itens, %d docs, %d fretes integrados"
          % (len(R["suppliers"]), len(R["product_items"]), len(R["docs"]), len(R["freights"])))


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    readme = build_readme()
    build_control_panel()
    build_calendar()
    build_products(); build_materials(); build_suppliers()
    build_bom(); build_bom_version()
    build_landed_cost(); build_tax_rules(); build_tax_credits()
    build_resource_rates(); build_standard_times()
    build_production_of(); build_actual_consumption()
    build_standard_cost(); build_actual_cost(); build_variances()
    build_capacity()
    build_structured_batch()
    build_cost_gap(); build_savings()
    build_cost_to_serve(); build_profitability(); build_pricing()
    build_recon_cost(); build_recon_accounting()
    build_dashboards()
    build_data_dictionary(); build_data_gaps(); build_legacy_recon()
    build_alerts()
    build_real_procurement()

    # fill README table of contents
    rr = readme._toc_start
    for name, grp, purp in TOC:
        put(readme, rr, 2, name, style="form", align="l")
        put(readme, rr, 3, grp, style="form", align="l")
        put(readme, rr, 4, purp, style="form", align="l")
        rr += 1

    # order sheets: README, CONTROL first; leave creation order otherwise
    wb.save(OUT)
    print("Saved %s with %d sheets" % (OUT, len(wb.sheetnames)))


if __name__ == "__main__":
    main()
