# -*- coding: utf-8 -*-
"""
Ingestao das extracoes fiscais reais da SKAL (notas de compra por mes).

Le todos os arquivos data/SKAL_Compras_*_Extracao*.xlsx (uma extracao por mes),
normaliza os cabecalhos (que variam entre meses) e devolve estruturas limpas:

  suppliers      : fornecedores deduplicados por CNPJ
  product_items  : itens de produto com quantidade E valor unitario numericos
                   (OCR sujo, sem numeros, e descartado — nunca inventamos)
  docs           : ledger documento-a-documento (Tributos_por_Doc): valor + ICMS
  freights       : CT-e de frete inbound
  energy         : itens de energia

Regra de ouro: campo nao legivel / nao numerico NAO e inferido.
"""
import glob
import os
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _num(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("R$", "").replace(" ", "")
    if not s:
        return None
    # brazilian 1.234,56 -> 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_name(s):
    if not s:
        return None
    s = re.sub(r"\s+", " ", str(s)).strip()
    # descarta lixo de OCR (linhas com pouca letra ou marcadores de rodape da SEFAZ)
    if len(s) < 4:
        return None
    low = s.lower()
    if "fazenda.gov" in low or "sefaz" in low or "portal nacional" in low:
        return None
    letters = sum(c.isalpha() for c in s)
    if letters < max(6, 0.45 * len(s)):
        return None
    # remove CNPJ-numero prefixado a alguns nomes de MEI
    s = re.sub(r"^\d{2}\.\d{3}\.\d{3}\s+", "", s)
    return s.upper()


def _month_from_name(path):
    m = re.search(r"(\d{4})-(\d{2})", os.path.basename(path))
    if m:
        return "%s-%s" % (m.group(1), m.group(2))
    for mes, num in [("Janeiro", "01"), ("Fevereiro", "02"), ("Marco", "03"),
                     ("Abril", "04"), ("Maio", "05"), ("Junho", "06"),
                     ("Julho", "07"), ("Agosto", "08"), ("Setembro", "09"),
                     ("Outubro", "10"), ("Novembro", "11"), ("Dezembro", "12")]:
        if mes.lower() in os.path.basename(path).lower():
            return "2026-" + num
    return "?"


def _rows(ws):
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hdr = [str(h).strip() if h is not None else "" for h in hdr]
    out = []
    for r in range(2, ws.max_row + 1):
        d = {hdr[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        out.append(d)
    return out


def load_all(data_dir="data"):
    result = {"suppliers": [], "product_items": [], "docs": [],
              "freights": [], "energy": [], "files": []}
    if openpyxl is None:
        return result
    seen_cnpj = {}
    for f in sorted(glob.glob(os.path.join(data_dir, "SKAL_Compras_*.xlsx"))):
        month = _month_from_name(f)
        result["files"].append((os.path.basename(f), month))
        wb = openpyxl.load_workbook(f, data_only=True)

        # ---- Fornecedores ----
        if "Fornecedores" in wb.sheetnames:
            for d in _rows(wb["Fornecedores"]):
                cnpj = (str(d.get("CNPJ")).strip() if d.get("CNPJ") else None)
                name = _clean_name(d.get("Razao_Social"))
                if not name:
                    continue
                key = cnpj or name
                tot = _num(d.get("Valor_Total")) or 0.0
                if key in seen_cnpj:
                    seen_cnpj[key]["total"] += tot
                    seen_cnpj[key]["months"].add(month)
                else:
                    rec = {"cnpj": cnpj, "name": name, "uf": d.get("UF"),
                           "regime": d.get("Regime"), "total": tot,
                           "months": {month}}
                    seen_cnpj[key] = rec
                    result["suppliers"].append(rec)

        # ---- Itens_Produtos (apenas linhas numericas) ----
        if "Itens_Produtos" in wb.sheetnames:
            for d in _rows(wb["Itens_Produtos"]):
                q = _num(d.get("Quantidade"))
                vu = _num(d.get("Valor_Unitario"))
                if not q or not vu:
                    continue
                desc = d.get("Descricao") or d.get("Descricao_Produto_OCR")
                emit = d.get("Emitente") or d.get("Fornecedor")
                result["product_items"].append({
                    "month": month, "doc": d.get("Doc_ID"),
                    "emit": _clean_name(emit) or (str(emit) if emit else None),
                    "desc": re.sub(r"\s+", " ", str(desc)).strip() if desc else None,
                    "ncm": d.get("NCM_SH"), "un": d.get("Unidade"),
                    "qty": q, "vunit": vu,
                    "vtot": _num(d.get("Valor_Total")) or _num(d.get("Valor_Produtos")) or round(q * vu, 2),
                    "icms": _num(d.get("ICMS")),
                })

        # ---- Tributos_por_Doc (ledger de documentos) ----
        if "Tributos_por_Doc" in wb.sheetnames:
            for d in _rows(wb["Tributos_por_Doc"]):
                val = _num(d.get("Valor_Documento"))
                emit = _clean_name(d.get("Emitente"))
                if not val and not emit:
                    continue
                result["docs"].append({
                    "month": month, "doc": d.get("Doc_ID"), "tipo": d.get("Tipo"),
                    "num": d.get("Numero"), "unidade": d.get("Unidade"),
                    "emit": emit or (str(d.get("Emitente")) if d.get("Emitente") else None),
                    "cnpj": d.get("CNPJ_Emitente"), "valor": val,
                    "bc_icms": _num(d.get("BC_ICMS")), "aliq_icms": _num(d.get("Aliq_ICMS")),
                    "icms": _num(d.get("ICMS")),
                })

        # ---- Fretes_CTe ----
        if "Fretes_CTe" in wb.sheetnames:
            for d in _rows(wb["Fretes_CTe"]):
                transp = _clean_name(d.get("Transportadora"))
                if not transp:
                    continue
                val = None
                for k in d:
                    if "valor" in str(k).lower():
                        val = _num(d.get(k)) or val
                result["freights"].append({
                    "month": month, "cte": d.get("CTe"), "transp": transp,
                    "origem": d.get("Origem"), "destino": d.get("Destino"),
                    "remetente": _clean_name(d.get("Remetente")), "valor": val,
                    "nf_rel": d.get("NF_e_Relacionada"),
                })

        # ---- Energia ----
        if "Energia" in wb.sheetnames:
            for d in _rows(wb["Energia"]):
                val = _num(d.get("Valor"))
                if not val:
                    continue
                result["energy"].append({
                    "month": month, "doc": d.get("Doc_ID"), "item": d.get("Item"),
                    "mwh": _num(d.get("Quantidade_MWh")),
                    "vunit": _num(d.get("Valor_Unitario_R_MWh")),
                    "valor": val, "icms_st": _num(d.get("ICMS_ST")),
                })

    result["suppliers"].sort(key=lambda x: -x["total"])
    return result


if __name__ == "__main__":
    r = load_all()
    print("files:", r["files"])
    print("suppliers:", len(r["suppliers"]), "| product_items:", len(r["product_items"]),
          "| docs:", len(r["docs"]), "| freights:", len(r["freights"]), "| energy:", len(r["energy"]))
    print("\nTOP 10 fornecedores por gasto:")
    for s in r["suppliers"][:10]:
        print("  R$ %10.2f  %-42s %s  %s" % (s["total"], s["name"][:42], s["uf"] or "--", sorted(s["months"])))
    print("\nITENS DE PRODUTO (limpos):")
    for it in r["product_items"]:
        print("  %s %-38s q=%.2f %s vu=%.2f tot=%.2f" % (it["month"], (it["desc"] or "")[:38], it["qty"], it["un"], it["vunit"], it["vtot"]))
    print("\nFRETES:", [(f["month"], f["cte"], f["transp"][:20], f["valor"]) for f in r["freights"]])
